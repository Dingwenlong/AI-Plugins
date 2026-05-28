#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from jsonschema import Draft202012Validator
except ModuleNotFoundError:
    Draft202012Validator = None

from chain_workspace import update_chain_status
from runtime import (
    COMPATIBLE_UPSTREAM_API_SPEC_SCHEMA_VERSIONS,
    SKILL_NAME,
    STATE_SCHEMA_VERSION,
    ExecutionContext,
    append_progress,
    configure_stdio,
    default_batch_file,
    dump_json,
    dump_text,
    load_json,
    load_batch_file,
    load_schema,
    normalize_persisted_path,
    now_iso,
    remove_file,
    resolve_agent_dir,
    resolve_context_root,
    resolve_project_root,
    resolve_solution_path,
    save_batch_file,
)


FIXTURE_STATUSES = {"pending", "in_progress", "done", "skipped", "not_required", "blocked", "error"}
TERMINAL_FIXTURE_STATUSES = {"done", "skipped", "not_required"}
UPSTREAM_READY_STATUS = "done"
DEFAULT_SQLSERVER_DATABASE = ""
APP_SAMPLE_SQL_LITERAL = "'APP'"
MISSING_SCHEMA_AUTHORITY_REASON = "missing_schema_authority"
UNSAFE_DATABASE_TARGET_REASON = "unsafe_database_target"
MISSING_DB_TARGET_REASON = "missing_db_target"
SQLITE_TARGET_DISABLED_REASON = "sqlite_target_disabled"
SQL_FIXTURE_TARGETS_LOCAL_CONFIG = Path("config") / "sql-fixture-targets.local.json"
SAFE_SQL_FIXTURE_ENVIRONMENTS = {"local", "test", "fixture", "sandbox", "integration", "develop"}
SQLSERVER_ALLOWED_TARGET_RULES: list[Any] = []
SQLSERVER_ALLOWED_SERVERS: set[str] = set()
SQLSERVER_ALLOWED_DATABASES: set[str] = set()
SQLSERVER_ALLOWED_SOURCES: set[str] = set()
SQLSERVER_ALLOWED_CONNECTION_NAMES: set[str] = set()
SQL_TABLE_PATTERN = re.compile(
    r"(?is)\b(?:from|join|into|update)\s+((?:\[[^\]]+\]|[A-Za-z_][A-Za-z0-9_$]*)"
    r"(?:\s*\.\s*(?:\[[^\]]+\]|[A-Za-z_][A-Za-z0-9_$]*)){0,2})"
)


class SkillError(RuntimeError):
    def __init__(self, message: str, *, status: str = "error", diagnosis_type: str | None = None) -> None:
        super().__init__(message)
        self.status = status
        self.diagnosis_type = diagnosis_type


@dataclass(frozen=True)
class SimpleValidationError:
    absolute_path: tuple[object, ...]
    message: str


class ZhArgumentParser(argparse.ArgumentParser):
    def __init__(self, *args: object, **kwargs: object) -> None:
        kwargs.setdefault("add_help", False)
        super().__init__(*args, **kwargs)
        self._positionals.title = "位置参数"
        self._optionals.title = "可选参数"

    def format_usage(self) -> str:
        return super().format_usage().replace("usage:", "用法：", 1)

    def format_help(self) -> str:
        text = super().format_help()
        return (
            text.replace("usage:", "用法：", 1)
            .replace("positional arguments:", "位置参数：")
            .replace("optional arguments:", "可选参数：")
            .replace("options:", "可选参数：")
            .replace("show this help message and exit", "显示此帮助并退出")
        )

    def error(self, message: str) -> None:
        self.print_usage(sys.stderr)
        self.exit(2, f"{self.prog}: 错误：{message}\n")


@dataclass(frozen=True)
class UpstreamApiRecord:
    api_id: str
    api_category: str
    api_name: str
    status: str
    block_reason: str | None
    manifest_path: Path
    api_spec_path: Path | None
    manifest_payload: dict[str, Any]
    api_spec_payload: dict[str, Any] | None


@dataclass(frozen=True)
class SqlServerTarget:
    server: str
    database: str
    integrated_security: bool
    username: str | None
    password: str | None
    trust_server_certificate: bool
    source: str
    connection_name: str | None = None
    target_name: str | None = None
    environment: str | None = None
    allow_create_table: bool = False
    allow_seed: bool = False


@dataclass(frozen=True)
class BlockedDbTarget:
    raw_target: str | None
    block_reason: str
    message: str


@dataclass(frozen=True)
class TableRefParts:
    database: str | None
    schema: str | None
    table: str


def parse_args() -> argparse.Namespace:
    parser = ZhArgumentParser(description="读取共享 .agent/context execution，检查 SQL fixture 需求并回写 fixture* 状态。")
    parser.add_argument("-h", "--help", action="help", help="显示此帮助并退出")
    parser.add_argument("--project-root", required=True)
    parser.add_argument("--solution-path", default=None)
    parser.add_argument("--agent-dir", default=".agent")
    parser.add_argument("--agent-root", help="集中 .agent 根目录；优先级高于环境变量与插件本地配置。")
    parser.add_argument("--workspace-root", help="共享工作区根目录，例如 D:\\Repo\\Project。")
    parser.add_argument("--workspace-key", help="插件 local-workspaces.json 中的工作区 key，例如 PROJECT。")
    parser.add_argument("--rules-root", help="专案规则库根目录；优先级高于环境变量与 workspace 配置。")
    parser.add_argument("--context-root", default=None)
    parser.add_argument("--function-code", default=None)
    parser.add_argument("--api-id", default=None)
    parser.add_argument("--execution-mode", choices=["auto", "prepare", "apply"], default="auto")
    parser.add_argument("--db-target", default=None)
    parser.add_argument("--schema-authority-root", default=None)
    parser.add_argument("--allow-create-database", action="store_true")
    parser.add_argument("--allow-create-table", action="store_true")
    parser.add_argument("--allow-seed", action="store_true")
    return parser.parse_args()


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    lines = [segment.strip() for segment in text.split("\n")]
    return "\n".join(segment for segment in lines if segment)


def format_validation_path(path_items: list[object] | tuple[object, ...]) -> str:
    rendered = "".join(
        f"[{item}]" if isinstance(item, int) else (f".{item}" if index else str(item))
        for index, item in enumerate(path_items)
    )
    return rendered or "$"


def schema_type_matches(value: Any, expected_type: str) -> bool:
    if expected_type == "object":
        return isinstance(value, dict)
    if expected_type == "array":
        return isinstance(value, list)
    if expected_type == "string":
        return isinstance(value, str)
    if expected_type == "integer":
        return isinstance(value, int) and not isinstance(value, bool)
    if expected_type == "number":
        return isinstance(value, (int, float)) and not isinstance(value, bool)
    if expected_type == "boolean":
        return isinstance(value, bool)
    if expected_type == "null":
        return value is None
    return True


def simple_schema_errors(value: Any, schema: dict[str, Any], path: tuple[object, ...] = ()) -> list[SimpleValidationError]:
    errors: list[SimpleValidationError] = []
    expected_type = schema.get("type")
    if expected_type is not None:
        expected_types = expected_type if isinstance(expected_type, list) else [expected_type]
        if not any(schema_type_matches(value, clean_text(item)) for item in expected_types):
            rendered = " or ".join(clean_text(item) for item in expected_types)
            return [SimpleValidationError(path, f"{value!r} is not of type {rendered}")]

    if "const" in schema and value != schema["const"]:
        errors.append(SimpleValidationError(path, f"{value!r} was expected to be constant {schema['const']!r}"))
    if "enum" in schema and value not in list(schema.get("enum") or []):
        errors.append(SimpleValidationError(path, f"{value!r} is not one of {list(schema.get('enum') or [])!r}"))

    if isinstance(value, dict):
        for required_key in list(schema.get("required") or []):
            if required_key not in value:
                errors.append(SimpleValidationError((*path, required_key), "is a required property"))
        properties = schema.get("properties") if isinstance(schema.get("properties"), dict) else {}
        for key, child_schema in properties.items():
            if key in value and isinstance(child_schema, dict):
                errors.extend(simple_schema_errors(value[key], child_schema, (*path, key)))
    elif isinstance(value, list):
        item_schema = schema.get("items")
        if isinstance(item_schema, dict):
            for index, item in enumerate(value):
                errors.extend(simple_schema_errors(item, item_schema, (*path, index)))
    return errors


def validate_payload_against_schema(payload: dict[str, Any], schema_name: str, label: str) -> None:
    schema = load_schema(schema_name)
    if Draft202012Validator is None:
        errors = simple_schema_errors(payload, schema)
    else:
        validator = Draft202012Validator(schema)
        errors = sorted(validator.iter_errors(payload), key=lambda item: list(item.absolute_path))
    if errors:
        first = errors[0]
        raise SkillError(f"{label} schema 校验失败：{format_validation_path(list(first.absolute_path))} - {first.message}")


def require_json_object(label: str, path: Path) -> dict[str, Any]:
    payload = load_json(path)
    if not isinstance(payload, dict):
        raise SkillError(f"{label} must be a JSON object: {path.as_posix()}")
    return payload


def normalize_identifier(value: str) -> str:
    text = clean_text(value).replace("[", "").replace("]", "").replace('"', "").replace("`", "")
    text = re.sub(r"\s+", "", text)
    return text


def normalize_slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", clean_text(value).casefold()).strip("_")


def leaf_table_name(value: str) -> str:
    normalized = normalize_identifier(value)
    return normalized.split(".")[-1] if normalized else ""


def parse_table_ref(value: str) -> TableRefParts:
    normalized = normalize_identifier(value)
    parts = [segment for segment in normalized.split(".") if segment]
    if len(parts) >= 3:
        return TableRefParts(database=parts[-3], schema=parts[-2], table=parts[-1])
    if len(parts) == 2:
        return TableRefParts(database=None, schema=parts[0], table=parts[1])
    return TableRefParts(database=None, schema=None, table=parts[0] if parts else normalized)


def dedupe_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        text = clean_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
    return ordered


def resolve_execution_root(context_root: Path, requested_function_code: str | None) -> tuple[Path, str, Path]:
    batch_file = default_batch_file(context_root)
    batch_payload = load_batch_file(batch_file)
    function_code = clean_text(requested_function_code) or clean_text(batch_payload.get("activeFunctionCode"))
    if function_code:
        execution_root = (context_root / function_code).resolve()
        if not execution_root.exists() or not execution_root.is_dir():
            raise SkillError(f"functionCode not found under context root: {function_code}")
        return execution_root, function_code, batch_file

    candidates = sorted(
        path.resolve()
        for path in context_root.iterdir()
        if path.is_dir() and (path / "execution-state.json").exists() and (path / "api-checklist.json").exists()
    )
    if not candidates:
        raise SkillError("context 目录下未找到可用 execution。")
    if len(candidates) > 1:
        rendered = "\n".join(f"- {path.name}" for path in candidates)
        raise SkillError(f"检测到多个 execution，请显式提供 --function-code 或设置 execution-batch.json.activeFunctionCode：\n{rendered}")
    return candidates[0], candidates[0].name, batch_file


def resolve_api_spec_path(api_dir: Path, api_id: str) -> Path | None:
    candidates = sorted(path.resolve() for path in api_dir.glob("*_API_Spec.json") if path.is_file())
    if not candidates:
        return None
    if len(candidates) > 1:
        raise SkillError(f"multiple API_Spec.json files found for upstream apiId: {api_id}")
    return candidates[0]


def resolve_context_rules_root(agent_dir: Path) -> Path | None:
    snapshot_path = agent_dir / "config" / "chain-workspace.json"
    if not snapshot_path.exists():
        return None
    try:
        payload = json.loads(snapshot_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return None
    raw_path = clean_text(payload.get("rulesRoot")) if isinstance(payload, dict) else ""
    if not raw_path:
        return None
    path = Path(raw_path).expanduser()
    if not path.is_absolute():
        path = agent_dir / path
    return path.resolve()


def reset_project_sql_defaults() -> None:
    global DEFAULT_SQLSERVER_DATABASE, APP_SAMPLE_SQL_LITERAL
    global SQLSERVER_ALLOWED_TARGET_RULES, SQLSERVER_ALLOWED_SERVERS
    global SQLSERVER_ALLOWED_DATABASES, SQLSERVER_ALLOWED_SOURCES, SQLSERVER_ALLOWED_CONNECTION_NAMES
    DEFAULT_SQLSERVER_DATABASE = ""
    APP_SAMPLE_SQL_LITERAL = "'APP'"
    SQLSERVER_ALLOWED_TARGET_RULES = []
    SQLSERVER_ALLOWED_SERVERS = set()
    SQLSERVER_ALLOWED_DATABASES = set()
    SQLSERVER_ALLOWED_SOURCES = set()
    SQLSERVER_ALLOWED_CONNECTION_NAMES = set()


def list_from_rule_value(value: Any) -> list[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def collect_rule_values(defaults: dict[str, Any], keys: tuple[str, ...]) -> list[Any]:
    values: list[Any] = []
    for key in keys:
        values.extend(list_from_rule_value(defaults.get(key)))
    return values


def normalized_rule_strings(defaults: dict[str, Any], keys: tuple[str, ...]) -> set[str]:
    return {clean_text(item).casefold() for item in collect_rule_values(defaults, keys) if clean_text(item)}


def configure_project_sql_defaults(agent_dir: Path) -> None:
    global DEFAULT_SQLSERVER_DATABASE, APP_SAMPLE_SQL_LITERAL
    global SQLSERVER_ALLOWED_TARGET_RULES, SQLSERVER_ALLOWED_SERVERS
    global SQLSERVER_ALLOWED_DATABASES, SQLSERVER_ALLOWED_SOURCES, SQLSERVER_ALLOWED_CONNECTION_NAMES
    reset_project_sql_defaults()
    rules_root = resolve_context_rules_root(agent_dir)
    if rules_root is None:
        return
    defaults_path = rules_root / "rules" / "sql-fixture" / "defaults.json"
    if not defaults_path.exists():
        return
    try:
        defaults = json.loads(defaults_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return
    if not isinstance(defaults, dict):
        return
    database = clean_text(defaults.get("defaultSqlServerDatabase"))
    app_literal = clean_text(defaults.get("appSampleSqlLiteral"))
    if database:
        DEFAULT_SQLSERVER_DATABASE = database
    if app_literal:
        APP_SAMPLE_SQL_LITERAL = app_literal
    SQLSERVER_ALLOWED_TARGET_RULES = collect_rule_values(
        defaults,
        (
            "allowedSqlServerTargets",
            "safeSqlServerTargets",
            "sqlServerTargetAllowlist",
            "sqlServerFixtureTargetAllowlist",
        ),
    )
    SQLSERVER_ALLOWED_SERVERS = normalized_rule_strings(defaults, ("allowedSqlServerServers", "safeSqlServerServers"))
    SQLSERVER_ALLOWED_DATABASES = normalized_rule_strings(defaults, ("allowedSqlServerDatabases", "safeSqlServerDatabases"))
    SQLSERVER_ALLOWED_SOURCES = normalized_rule_strings(defaults, ("allowedSqlServerSources", "safeSqlServerSources"))
    SQLSERVER_ALLOWED_CONNECTION_NAMES = normalized_rule_strings(
        defaults,
        ("allowedSqlServerConnectionNames", "safeSqlServerConnectionNames"),
    )


def build_context(args: argparse.Namespace) -> tuple[ExecutionContext, dict[str, Any], dict[str, Any], dict[str, UpstreamApiRecord]]:
    project_root = resolve_project_root(args.project_root)
    agent_dir = resolve_agent_dir(project_root, args.agent_dir, args.agent_root, args.workspace_root, args.workspace_key, args.rules_root)
    configure_project_sql_defaults(agent_dir)
    solution_path = resolve_solution_path(project_root, args.solution_path)
    context_root = resolve_context_root(project_root, agent_dir, args.context_root)
    execution_root, execution_id, batch_file = resolve_execution_root(context_root, args.function_code)
    execution_state = require_json_object("execution-state.json", execution_root / "execution-state.json")
    checklist_payload = require_json_object("api-checklist.json", execution_root / "api-checklist.json")
    items = checklist_payload.get("items")
    if not isinstance(items, list):
        raise SkillError("api-checklist.json.items must be an array.")
    api_map: dict[str, UpstreamApiRecord] = {}
    for raw_item in items:
        if not isinstance(raw_item, dict):
            raise SkillError("api-checklist.json.items must contain objects.")
        api_id = clean_text(raw_item.get("apiId"))
        spec_status = clean_text(raw_item.get("specStatus"))
        if not api_id:
            continue
        api_dir = execution_root / "apis" / api_id
        manifest_path = api_dir / "manifest.json"
        manifest_payload = require_json_object("manifest.json", manifest_path)
        api_spec_path = resolve_api_spec_path(api_dir, api_id) if spec_status == UPSTREAM_READY_STATUS else None
        api_spec_payload = require_json_object("upstream API_Spec.json", api_spec_path) if api_spec_path else None
        if api_spec_payload and clean_text(api_spec_payload.get("schemaVersion")) not in COMPATIBLE_UPSTREAM_API_SPEC_SCHEMA_VERSIONS:
            supported = ", ".join(sorted(COMPATIBLE_UPSTREAM_API_SPEC_SCHEMA_VERSIONS))
            raise SkillError(f"upstream API_Spec schemaVersion mismatch for {api_id}; supported: {supported}")
        api_map[api_id] = UpstreamApiRecord(
            api_id=api_id,
            api_category=clean_text(raw_item.get("apiCategory")),
            api_name=clean_text(raw_item.get("apiName")),
            status=spec_status,
            block_reason=clean_text(raw_item.get("specBlockReason")) or None,
            manifest_path=manifest_path.resolve(),
            api_spec_path=api_spec_path,
            manifest_payload=manifest_payload,
            api_spec_payload=api_spec_payload,
        )
    context = ExecutionContext(
        project_root=project_root,
        solution_path=solution_path,
        agent_dir=agent_dir,
        context_root=context_root,
        batch_file=batch_file,
        state_root=execution_root,
        execution_id=execution_id,
    )
    return context, execution_state, checklist_payload, api_map


def extract_query_texts(api_spec: dict[str, Any]) -> tuple[list[str], list[str]]:
    sources_used: list[str] = []
    query_texts: list[str] = []
    code_handoff = api_spec.get("codeHandoff")
    if isinstance(code_handoff, dict):
        for contract in list(code_handoff.get("queryContracts") or []):
            if isinstance(contract, dict) and clean_text(contract.get("sqlText")):
                query_texts.append(clean_text(contract.get("sqlText")))
                sources_used.append("codeHandoff.queryContracts")
    business_logic = api_spec.get("businessLogic")
    if isinstance(business_logic, dict):
        for spec in list(business_logic.get("sqlSpecs") or []):
            if not isinstance(spec, dict):
                continue
            sql_text = clean_text(spec.get("sqlText") or spec.get("queryText"))
            if sql_text:
                query_texts.append(sql_text)
                sources_used.append("businessLogic.sqlSpecs")
    return dedupe_strings(query_texts), dedupe_strings(sources_used)


def extract_backend_table_refs(api_spec: dict[str, Any]) -> tuple[list[str], list[str]]:
    backend_apis = api_spec.get("backendApis")
    if not isinstance(backend_apis, dict):
        return [], []
    refs: list[str] = []
    systems: list[str] = []
    for system, values in backend_apis.items():
        system_name = clean_text(system)
        candidates = values if isinstance(values, list) else [values]
        for candidate in candidates:
            text = clean_text(candidate)
            if not text:
                continue
            if "." in text and "(" not in text:
                refs.append(text)
                if system_name:
                    systems.append(system_name)
    return dedupe_strings(refs), dedupe_strings(systems)


def extract_tables_from_sql(query_texts: list[str]) -> list[str]:
    tables: list[str] = []
    for text in query_texts:
        for match in SQL_TABLE_PATTERN.finditer(text):
            candidate = normalize_identifier(match.group(1))
            if candidate:
                tables.append(candidate)
    return dedupe_strings(tables)


def detect_sql_fixture_need(api_spec: dict[str, Any]) -> dict[str, Any]:
    query_texts, query_sources = extract_query_texts(api_spec)
    backend_refs, backend_sources = extract_backend_table_refs(api_spec)
    table_refs = dedupe_strings(extract_tables_from_sql(query_texts) + backend_refs)
    sql_required = bool(query_texts or table_refs)
    return {
        "sqlFixtureRequired": sql_required,
        "sourcesUsed": dedupe_strings(query_sources + (["backendApis"] if backend_refs else [])),
        "queryTexts": query_texts,
        "tableRefs": table_refs,
        "backendSystems": backend_sources,
    }


def strip_json_comments(text: str) -> str:
    result: list[str] = []
    in_string = False
    escape = False
    index = 0
    while index < len(text):
        char = text[index]
        next_char = text[index + 1] if index + 1 < len(text) else ""
        if in_string:
            result.append(char)
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            index += 1
            continue
        if char == '"':
            in_string = True
            result.append(char)
            index += 1
            continue
        if char == "/" and next_char == "/":
            index += 2
            while index < len(text) and text[index] not in "\r\n":
                index += 1
            continue
        if char == "/" and next_char == "*":
            index += 2
            while index + 1 < len(text) and not (text[index] == "*" and text[index + 1] == "/"):
                index += 1
            index += 2
            continue
        result.append(char)
        index += 1
    return "".join(result)


def load_json_with_comments(path: Path) -> dict[str, Any]:
    payload = json.loads(strip_json_comments(path.read_text(encoding="utf-8-sig")))
    if not isinstance(payload, dict):
        raise SkillError(f"appsettings file must be a JSON object: {path.as_posix()}")
    return payload


def parse_ado_connection_string(connection_string: str) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for segment in connection_string.split(";"):
        if "=" not in segment:
            continue
        key, value = segment.split("=", 1)
        key_text = clean_text(key).casefold()
        value_text = value.strip().strip('"')
        if key_text:
            pairs[key_text] = value_text
    return pairs


def load_connection_strings_from_appsettings(project_root: Path) -> dict[str, str]:
    candidates = sorted(path for path in project_root.rglob("appsettings.json") if path.is_file())
    merged: dict[str, str] = {}
    preferred: list[Path] = []
    fallback: list[Path] = []
    for path in candidates:
        lowered = path.as_posix().casefold()
        if "/api/enterpriseapi/enterpriseapi/" in lowered:
            preferred.append(path)
        else:
            fallback.append(path)
    for path in [*preferred, *fallback]:
        try:
            payload = load_json_with_comments(path)
        except Exception:
            continue
        connection_strings = payload.get("ConnectionStrings")
        if not isinstance(connection_strings, dict):
            continue
        for key, value in connection_strings.items():
            key_text = clean_text(key)
            value_text = clean_text(value)
            if key_text and value_text and key_text not in merged:
                merged[key_text] = value_text
    return merged


def is_sqlserver_connection_string(connection_string: str) -> bool:
    text = clean_text(connection_string).casefold()
    return "data source=" in text or "server=" in text


def is_local_sqlserver_target(target: SqlServerTarget) -> bool:
    server = clean_text(target.server).casefold()
    return (
        server.startswith("(localdb)")
        or server in {".", "(local)", "localhost", "127.0.0.1"}
        or server.startswith(".\\")
        or server.startswith("localhost\\")
    )


def build_sqlserver_target_from_connection_string(
    connection_string: str,
    *,
    source: str,
    connection_name: str | None = None,
) -> SqlServerTarget:
    parsed = parse_ado_connection_string(connection_string)
    server = clean_text(parsed.get("data source") or parsed.get("server"))
    database = clean_text(parsed.get("initial catalog") or parsed.get("database"))
    integrated_security = clean_text(parsed.get("integrated security")).casefold() in {"true", "sspi", "yes"}
    username = clean_text(parsed.get("user id") or parsed.get("uid")) or None
    password = clean_text(parsed.get("password") or parsed.get("pwd")) or None
    trust_server_certificate = clean_text(parsed.get("trustservercertificate")).casefold() in {"true", "yes"}
    if not server or not database:
        raise SkillError(f"invalid SQL Server connection string: missing server/database from {source}")
    return SqlServerTarget(
        server=server,
        database=database,
        integrated_security=integrated_security,
        username=username,
        password=password,
        trust_server_certificate=trust_server_certificate,
        source=source,
        connection_name=connection_name,
    )


def redact_db_target_text(value: str | None) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    segments = []
    for segment in text.split(";"):
        if "=" not in segment:
            segments.append(segment)
            continue
        key, raw_value = segment.split("=", 1)
        if clean_text(key).casefold().endswith(("password", "pwd")):
            segments.append(f"{key}=***")
        else:
            segments.append(f"{key}={raw_value}")
    return ";".join(segments)


def sqlserver_source_keys(target: SqlServerTarget) -> set[str]:
    keys = {clean_text(target.source).casefold()}
    if target.connection_name:
        connection_name = clean_text(target.connection_name).casefold()
        keys.add(connection_name)
        keys.add(f"connectionstrings.{connection_name}")
    return {key for key in keys if key}


def sqlserver_target_signature(target: SqlServerTarget) -> str:
    return f"{clean_text(target.server).casefold()}|{clean_text(target.database).casefold()}"


def sqlserver_target_rule_matches(target: SqlServerTarget, rule: Any) -> bool:
    if isinstance(rule, str):
        text = clean_text(rule)
        lowered = text.casefold()
        if is_sqlserver_connection_string(text):
            try:
                rule_target = build_sqlserver_target_from_connection_string(text, source="project_rules")
            except SkillError:
                return False
            return sqlserver_target_signature(rule_target) == sqlserver_target_signature(target)
        return lowered in {
            *sqlserver_source_keys(target),
            clean_text(target.server).casefold(),
            clean_text(target.database).casefold(),
            sqlserver_target_signature(target),
        }

    if not isinstance(rule, dict):
        return False
    checks: list[bool] = []
    server = clean_text(rule.get("server") or rule.get("dataSource"))
    database = clean_text(rule.get("database") or rule.get("initialCatalog"))
    source = clean_text(rule.get("source"))
    connection_name = clean_text(rule.get("connectionName") or rule.get("connection"))
    connection_string = clean_text(rule.get("connectionString"))
    if server:
        checks.append(server.casefold() == clean_text(target.server).casefold())
    if database:
        checks.append(database.casefold() == clean_text(target.database).casefold())
    if source:
        checks.append(source.casefold() in sqlserver_source_keys(target))
    if connection_name:
        checks.append(connection_name.casefold() in sqlserver_source_keys(target))
    if connection_string:
        try:
            rule_target = build_sqlserver_target_from_connection_string(connection_string, source="project_rules")
        except SkillError:
            checks.append(False)
        else:
            checks.append(sqlserver_target_signature(rule_target) == sqlserver_target_signature(target))
    return bool(checks) and all(checks)


def sqlserver_target_safety_proof(target: SqlServerTarget) -> str | None:
    if target.target_name and clean_text(target.source).startswith(".agent/config/sql-fixture-targets.local.json"):
        return "agent_local_sql_fixture_target"
    return None


def require_safe_sqlserver_target(target: SqlServerTarget, raw_target: str | None = None) -> SqlServerTarget | BlockedDbTarget:
    if sqlserver_target_safety_proof(target):
        return target
    label = redact_db_target_text(raw_target) or target.source
    return BlockedDbTarget(
        raw_target=label,
        block_reason=UNSAFE_DATABASE_TARGET_REASON,
        message=f"SQL Server fixture target must be configured in .agent/config/sql-fixture-targets.local.json: {label}",
    )


def sql_fixture_target_config_path(agent_dir: Path) -> Path:
    return agent_dir / SQL_FIXTURE_TARGETS_LOCAL_CONFIG


def bool_from_config(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = clean_text(value).casefold()
    return text in {"1", "true", "yes", "y"}


def allowed_sqlserver_target_database_names() -> set[str]:
    names = set(SQLSERVER_ALLOWED_DATABASES)
    if clean_text(DEFAULT_SQLSERVER_DATABASE):
        names.add(clean_text(DEFAULT_SQLSERVER_DATABASE).casefold())
    return names


def validate_configured_target_database(target_database: str) -> str | BlockedDbTarget:
    database = clean_text(target_database)
    if not database:
        return BlockedDbTarget(
            raw_target=".agent/config/sql-fixture-targets.local.json",
            block_reason=MISSING_DB_TARGET_REASON,
            message="SQL fixture targetDatabase is required.",
        )
    allowed_databases = allowed_sqlserver_target_database_names()
    if allowed_databases and database.casefold() not in allowed_databases:
        return BlockedDbTarget(
            raw_target=f"targetDatabase={database}",
            block_reason=UNSAFE_DATABASE_TARGET_REASON,
            message=f"SQL fixture targetDatabase is not allowed by project rules: {database}",
        )
    return database


def read_sql_fixture_targets_config(agent_dir: Path) -> dict[str, Any] | None:
    config_path = sql_fixture_target_config_path(agent_dir)
    if not config_path.exists():
        return None
    try:
        payload = json.loads(config_path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise SkillError(
            f"Invalid SQL fixture target config: {config_path}",
            status="blocked",
            diagnosis_type="invalid_db_target_config",
        ) from exc
    if not isinstance(payload, dict):
        raise SkillError(
            f"SQL fixture target config must be a JSON object: {config_path}",
            status="blocked",
            diagnosis_type="invalid_db_target_config",
        )
    return payload


def configured_sqlserver_target_report_label(target: SqlServerTarget) -> str:
    pieces = [
        f"source={target.source}",
        f"target={target.target_name}" if target.target_name else "",
        f"environment={target.environment}" if target.environment else "",
        f"server={target.server}",
        f"database={target.database}",
    ]
    return ";".join(piece for piece in pieces if piece)


def select_sql_fixture_target_payload(payload: dict[str, Any], target_name: str | None) -> tuple[str, dict[str, Any] | None]:
    selected_name = clean_text(target_name) or clean_text(payload.get("defaultTarget"))
    targets = payload.get("targets")
    if isinstance(targets, dict):
        if not selected_name and len(targets) == 1:
            selected_name = clean_text(next(iter(targets.keys())))
        target_payload = targets.get(selected_name) if selected_name else None
        return selected_name, target_payload if isinstance(target_payload, dict) else None
    if not selected_name:
        selected_name = "default"
    return selected_name, payload


def build_sqlserver_target_from_agent_config(agent_dir: Path, target_name: str | None = None) -> SqlServerTarget | BlockedDbTarget | None:
    payload = read_sql_fixture_targets_config(agent_dir)
    if payload is None:
        return None
    selected_name, target_payload = select_sql_fixture_target_payload(payload, target_name)
    if target_payload is None:
        return BlockedDbTarget(
            raw_target=f"target={selected_name or '<missing>'}",
            block_reason=MISSING_DB_TARGET_REASON,
            message="SQL fixture target config does not contain the selected target.",
        )

    provider = clean_text(target_payload.get("provider")).casefold()
    if provider != "sqlserver":
        return BlockedDbTarget(
            raw_target=f"target={selected_name};provider={provider or '<missing>'}",
            block_reason=UNSAFE_DATABASE_TARGET_REASON,
            message="SQL fixture target provider must be sqlserver.",
        )

    environment = clean_text(target_payload.get("environment"))
    if environment.casefold() not in SAFE_SQL_FIXTURE_ENVIRONMENTS:
        return BlockedDbTarget(
            raw_target=f"target={selected_name};environment={environment or '<missing>'}",
            block_reason=UNSAFE_DATABASE_TARGET_REASON,
            message="SQL fixture target environment must be local/test/fixture/sandbox/integration/develop.",
        )

    target_database = validate_configured_target_database(clean_text(target_payload.get("targetDatabase")))
    if isinstance(target_database, BlockedDbTarget):
        return target_database

    connection_string = clean_text(target_payload.get("connectionString"))
    if not connection_string:
        return BlockedDbTarget(
            raw_target=f"target={selected_name};connectionString=<missing>",
            block_reason=MISSING_DB_TARGET_REASON,
            message="SQL fixture target connectionString is required.",
        )
    if not is_sqlserver_connection_string(connection_string):
        return BlockedDbTarget(
            raw_target=f"target={selected_name};connectionString=<invalid>",
            block_reason=UNSAFE_DATABASE_TARGET_REASON,
            message="SQL fixture target connectionString must be a SQL Server connection string.",
        )

    parsed = build_sqlserver_target_from_connection_string(
        connection_string,
        source=f".agent/config/sql-fixture-targets.local.json:{selected_name}",
        connection_name=selected_name,
    )
    configured = SqlServerTarget(
        server=parsed.server,
        database=target_database,
        integrated_security=parsed.integrated_security,
        username=parsed.username,
        password=parsed.password,
        trust_server_certificate=parsed.trust_server_certificate,
        source=f".agent/config/sql-fixture-targets.local.json:{selected_name}",
        connection_name=selected_name,
        target_name=selected_name,
        environment=environment,
        allow_create_table=bool_from_config(target_payload.get("allowCreateTable")),
        allow_seed=bool_from_config(target_payload.get("allowSeed")),
    )
    return require_safe_sqlserver_target(configured, raw_target=configured_sqlserver_target_report_label(configured))


def blocked_explicit_db_target(raw_target: str, reason: str = UNSAFE_DATABASE_TARGET_REASON) -> BlockedDbTarget:
    label = redact_db_target_text(raw_target) or "<empty>"
    message = "SQLite fixture targets are disabled; configure SQL Server in .agent/config/sql-fixture-targets.local.json."
    if reason != SQLITE_TARGET_DISABLED_REASON:
        message = "SQL fixture targets must come from .agent/config/sql-fixture-targets.local.json."
    return BlockedDbTarget(raw_target=label, block_reason=reason, message=message)


def parse_db_target(
    db_target: str | None,
    project_root: Path,
    *,
    agent_dir: Path,
    preferred_connection_names: list[str] | None = None,
) -> SqlServerTarget | BlockedDbTarget | None:
    _ = project_root
    _ = preferred_connection_names
    text = clean_text(db_target)
    if not text:
        return build_sqlserver_target_from_agent_config(agent_dir)
    if text.casefold().startswith("agent-config:"):
        return build_sqlserver_target_from_agent_config(agent_dir, text.split(":", 1)[1].strip())
    if text.casefold().startswith("sqlite:///"):
        return blocked_explicit_db_target("sqlite:///<disabled>", SQLITE_TARGET_DISABLED_REASON)
    if text.casefold().startswith("connection-name:"):
        return blocked_explicit_db_target(text)
    if text.casefold().startswith("connstr:"):
        return blocked_explicit_db_target(text)
    if "data source=" in text.casefold() or "server=" in text.casefold():
        return blocked_explicit_db_target(text)
    raise SkillError(f"unsupported db target in skeleton: {text}", status="blocked", diagnosis_type="unsupported_db_target")


def run_sqlcmd(target: SqlServerTarget, query: str, *, database_override: str | None = None) -> subprocess.CompletedProcess[str]:
    database = database_override or target.database
    command = ["sqlcmd", "-S", target.server, "-d", database, "-Q", query, "-h", "-1", "-W", "-b"]
    if target.integrated_security:
        command.append("-E")
    else:
        if target.username is None or target.password is None:
            raise SkillError("SQL Server connection string is missing username/password.", status="blocked", diagnosis_type="invalid_connection_string")
        command.extend(["-U", target.username, "-P", target.password])
    if target.trust_server_certificate:
        command.append("-C")
    return subprocess.run(command, capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=30, check=False)


def sqlserver_database_exists(target: SqlServerTarget, database_name: str | None = None) -> bool:
    database = clean_text(database_name) or target.database
    escaped_name = database.replace("'", "''")
    result = run_sqlcmd(target, f"SET NOCOUNT ON; SELECT name FROM sys.databases WHERE name = '{escaped_name}';", database_override="master")
    if result.returncode != 0:
        raise SkillError(clean_text(result.stderr) or "sqlcmd failed while checking database existence.", status="blocked", diagnosis_type="sqlcmd_failed")
    return clean_text(result.stdout).casefold() == database.casefold()


def sqlserver_target_database(target: SqlServerTarget) -> str:
    return clean_text(DEFAULT_SQLSERVER_DATABASE) or target.database


def sqlserver_target_table_parts(target: SqlServerTarget, table_ref: str) -> TableRefParts:
    target_database = sqlserver_target_database(target)
    parts = parse_table_ref(table_ref)
    schema = parts.schema or "dbo"
    if parts.database:
        schema = parts.schema or "dbo"
    elif schema.casefold() == target_database.casefold():
        schema = "dbo"
    return TableRefParts(database=target_database, schema=schema, table=parts.table)


def sqlserver_find_existing_table(target: SqlServerTarget, table_ref: str) -> str | None:
    parts = sqlserver_target_table_parts(target, table_ref)
    database = parts.database or target.database
    schema_name = parts.schema.replace("'", "''") if parts.schema else None
    schema_predicate = f" AND TABLE_SCHEMA = '{schema_name}'" if schema_name else ""
    leaf = parts.table.replace("'", "''")
    result = run_sqlcmd(
        target,
        "SET NOCOUNT ON; "
        "SELECT TOP 1 TABLE_SCHEMA + '.' + TABLE_NAME "
        f"FROM [{database}].INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{leaf}'{schema_predicate} ORDER BY TABLE_SCHEMA, TABLE_NAME;",
        database_override="master",
    )
    if result.returncode != 0:
        raise SkillError(clean_text(result.stderr) or "sqlcmd failed while checking table existence.", status="blocked", diagnosis_type="sqlcmd_failed")
    value = clean_text(result.stdout)
    if value:
        return f"{database}.{value}"
    return None


def sqlserver_row_count(target: SqlServerTarget, table_ref: str, resolved_table_name: str) -> int:
    base_parts = parse_table_ref(table_ref)
    resolved_parts = [segment for segment in normalize_identifier(resolved_table_name).split(".") if segment]
    if len(resolved_parts) >= 3:
        database = resolved_parts[-3]
    else:
        database = base_parts.database or DEFAULT_SQLSERVER_DATABASE or target.database
    if len(resolved_parts) == 1:
        quoted_name = f"[dbo].[{resolved_parts[0]}]"
    else:
        quoted_name = ".".join(f"[{part}]" for part in resolved_parts[-2:])
    result = run_sqlcmd(target, f"SET NOCOUNT ON; SELECT COUNT(1) FROM [{database}].{quoted_name};", database_override="master")
    if result.returncode != 0:
        raise SkillError(clean_text(result.stderr) or "sqlcmd failed while counting rows.", status="blocked", diagnosis_type="sqlcmd_failed")
    value = clean_text(result.stdout)
    return int(value) if value.isdigit() else 0


def sqlserver_create_database(target: SqlServerTarget, database_name: str) -> None:
    escaped_name = database_name.replace("'", "''")
    result = run_sqlcmd(target, f"IF DB_ID(N'{escaped_name}') IS NULL CREATE DATABASE [{database_name}];", database_override="master")
    if result.returncode != 0:
        raise SkillError(clean_text(result.stderr) or "sqlcmd failed while creating database.", status="blocked", diagnosis_type="sqlcmd_failed")


def find_authority_root(project_root: Path, explicit_root: str | None) -> Path | None:
    if explicit_root:
        root = Path(explicit_root).expanduser()
        if not root.is_absolute():
            root = (project_root / root).resolve()
        return root if root.exists() else None
    default_root = (project_root / ".agent" / "Reference").resolve()
    return default_root if default_root.exists() else None


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def find_schema_sql(authority_root: Path | None, table_ref: str) -> Path | None:
    if authority_root is None:
        return None
    leaf = normalize_slug(leaf_table_name(table_ref))
    full = normalize_slug(normalize_identifier(table_ref).replace(".", "_"))
    candidates = {f"{leaf}.sql", f"{leaf}_schema.sql", f"{full}.sql", f"{full}_schema.sql"}
    lowered = {candidate.casefold() for candidate in candidates}
    for path in authority_root.rglob("*.sql"):
        if path.name.casefold() in lowered:
            return path.resolve()
    return None


def find_seed_sql(authority_root: Path | None, api_id: str, table_ref: str) -> Path | None:
    if authority_root is None:
        return None
    api_slug = normalize_slug(api_id)
    leaf = normalize_slug(leaf_table_name(table_ref))
    candidates = {
        f"{api_slug}.seed.sql",
        f"{api_slug}_seed.sql",
        f"{leaf}.seed.sql",
        f"{leaf}_seed.sql",
    }
    lowered = {candidate.casefold() for candidate in candidates}
    for path in authority_root.rglob("*.sql"):
        if path.name.casefold() in lowered:
            return path.resolve()
    return None


def load_db_schema_index(authority_root: Path | None) -> dict[str, Any] | None:
    if authority_root is None:
        return None
    index_path = authority_root / "indexes" / "db-schema-index.json"
    if not index_path.exists():
        return None
    payload = load_json(index_path)
    return payload if isinstance(payload, dict) else None


def find_workbook_sheet_for_table(project_root: Path, authority_root: Path | None, table_ref: str) -> tuple[Path, str] | None:
    index_payload = load_db_schema_index(authority_root)
    if not index_payload:
        index_items: list[dict[str, Any]] = []
    else:
        index_items = [item for item in list(index_payload.get("items") or []) if isinstance(item, dict)]
    table_name = parse_table_ref(table_ref).table
    for item in index_items:
        sheet_match_keys = item.get("sheetMatchKeys")
        if not isinstance(sheet_match_keys, dict):
            continue
        locator = sheet_match_keys.get(table_name)
        if not isinstance(locator, list) or not locator:
            continue
        sheet_name = clean_text(locator[0])
        relative_path = clean_text(item.get("relativePath"))
        if not sheet_name or not relative_path:
            continue
        workbook_path = (project_root / relative_path).resolve()
        if workbook_path.exists():
            return workbook_path, sheet_name
    if authority_root is None:
        return None
    for workbook_path in authority_root.rglob("*.xlsx"):
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        except Exception:
            continue
        for sheet_name in workbook.sheetnames:
            normalized_sheet = clean_text(sheet_name).upper()
            if normalized_sheet == table_name.upper() or normalized_sheet.startswith(f"{table_name.upper()}("):
                return workbook_path.resolve(), sheet_name
    return None


def load_authority_columns(project_root: Path, authority_root: Path | None, table_ref: str) -> list[dict[str, Any]]:
    workbook_sheet = find_workbook_sheet_for_table(project_root, authority_root, table_ref)
    if workbook_sheet is None:
        return []
    workbook_path, sheet_name = workbook_sheet
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_mode = None
    header_map: dict[str, int] = {}
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        values = [clean_text(cell) for cell in row]
        if "資料行名稱" in values and "資料類型" in values:
            header_mode = "app_table_schema"
            header_map = {value: index for index, value in enumerate(values) if value}
            break
    columns: list[dict[str, Any]] = []
    if header_mode == "app_table_schema":
        field_index = header_map.get("資料行名稱", 0)
        type_index = header_map.get("資料類型", 1)
        nullable_index = header_map.get("可為NULL", 2)
        for row in sheet.iter_rows(min_row=7, values_only=True):
            field_name = clean_text(row[field_index] if len(row) > field_index else None)
            data_type = clean_text(row[type_index] if len(row) > type_index else None)
            if not field_name and not data_type:
                if columns:
                    break
                continue
            if not field_name or not data_type:
                continue
            nullable_marker = clean_text(row[nullable_index] if len(row) > nullable_index else None)
            columns.append(
                {
                    "fieldName": field_name,
                    "dataType": data_type,
                    "nullable": nullable_marker in {"是", "Y", "y", "true", "TRUE"},
                    "defaultValue": None,
                }
            )
        return columns
    for row in sheet.iter_rows(min_row=6, values_only=True):
        field_name = clean_text(row[1] if len(row) > 1 else None)
        data_type = clean_text(row[6] if len(row) > 6 else None)
        if not field_name or not data_type:
            continue
        nullable_marker = clean_text(row[5] if len(row) > 5 else None).casefold()
        default_value = clean_text(row[7] if len(row) > 7 else None)
        columns.append(
            {
                "fieldName": field_name,
                "dataType": data_type,
                "nullable": nullable_marker == "v",
                "defaultValue": default_value or None,
            }
        )
    return columns


def sql_literal_for_sample(column_name: str, data_type: str) -> str:
    column = clean_text(column_name).upper()
    lowered = clean_text(data_type).casefold()
    if "datetime" in lowered or "date" in lowered:
        return "'2026-04-15T10:00:00'"
    if any(token in lowered for token in ["decimal", "numeric", "int", "bigint", "smallint", "tinyint"]):
        return "0"
    if column.endswith("COUNTRY_ID") or column == "COUNTRY":
        return "'US'"
    if column.endswith("COUNTRY_NAME"):
        return "'美國'"
    if column == "CUSTID":
        return "'C123456789'"
    if column == "SOURCE":
        return "'DAGRAPHICS'"
    if column == "LOGINTYPE":
        return APP_SAMPLE_SQL_LITERAL
    if column == "IP":
        return "'10.11.2.3'"
    if column == "OS":
        return "'iOS 18.0.1'"
    if column == "DEVICE":
        return "'iPhone 15 Pro Max'"
    return "'FIXTURE'"


def build_generated_schema_sql(table_ref: str, columns: list[dict[str, Any]]) -> str:
    parts = parse_table_ref(table_ref)
    schema_name = "dbo" if parts.database is None and clean_text(parts.schema).casefold() == DEFAULT_SQLSERVER_DATABASE.casefold() else (parts.schema or "dbo")
    lines = [
        f"IF OBJECT_ID(N'[{schema_name}].[{parts.table}]', N'U') IS NULL",
        "BEGIN",
        f"    CREATE TABLE [{schema_name}].[{parts.table}] (",
    ]
    rendered_columns: list[str] = []
    for column in columns:
        null_clause = "NULL" if bool(column.get("nullable")) else "NOT NULL"
        default_value = clean_text(column.get("defaultValue"))
        default_clause = f" DEFAULT {default_value}" if default_value else ""
        rendered_columns.append(
            f"        [{clean_text(column['fieldName'])}] {clean_text(column['dataType'])}{default_clause} {null_clause}"
        )
    lines.append(",\n".join(rendered_columns))
    lines.extend(["    );", "END", ""])
    return "\n".join(lines)


def build_generated_seed_sql(table_ref: str, columns: list[dict[str, Any]]) -> str:
    parts = parse_table_ref(table_ref)
    schema_name = "dbo" if parts.database is None and clean_text(parts.schema).casefold() == DEFAULT_SQLSERVER_DATABASE.casefold() else (parts.schema or "dbo")
    first_column = clean_text(columns[0]["fieldName"])
    first_value = sql_literal_for_sample(first_column, clean_text(columns[0]["dataType"]))
    column_list = ", ".join(f"[{clean_text(column['fieldName'])}]" for column in columns)
    value_list = ", ".join(sql_literal_for_sample(clean_text(column["fieldName"]), clean_text(column["dataType"])) for column in columns)
    return (
        f"IF NOT EXISTS (SELECT 1 FROM [{schema_name}].[{parts.table}] WHERE [{first_column}] = {first_value})\n"
        "BEGIN\n"
        f"    INSERT INTO [{schema_name}].[{parts.table}] ({column_list})\n"
        f"    VALUES ({value_list});\n"
        "END\n"
    )


def build_blocked_target_table_checks(
    context: ExecutionContext,
    api_id: str,
    table_refs: list[str],
    blocked_target: BlockedDbTarget,
    authority_root: Path | None,
) -> list[dict[str, Any]]:
    refs = table_refs or ["<database-target>"]
    return [
        {
            "tableRef": table_ref,
            "resolvedTableName": None,
            "exists": False,
            "rowCount": None,
            "action": "blocked",
            "blockReason": blocked_target.block_reason,
            "schemaSql": normalize_persisted_path(find_schema_sql(authority_root, table_ref), project_root=context.project_root),
            "seedSql": normalize_persisted_path(find_seed_sql(authority_root, api_id, table_ref), project_root=context.project_root),
        }
        for table_ref in refs
    ]


def ensure_generated_authority_sql(
    context: ExecutionContext,
    api_id: str,
    table_ref: str,
    authority_root: Path | None,
) -> tuple[Path | None, Path | None]:
    columns = load_authority_columns(context.project_root, authority_root, table_ref)
    if not columns:
        return None, None
    slug = normalize_slug(table_ref.replace(".", "_"))
    schema_path = context.paths.api_dir(api_id) / f"generated-schema-{slug}.sql"
    seed_path = context.paths.api_dir(api_id) / f"generated-seed-{slug}.sql"
    dump_text(schema_path, build_generated_schema_sql(table_ref, columns))
    dump_text(seed_path, build_generated_seed_sql(table_ref, columns))
    return schema_path, seed_path


def inspect_sqlserver_tables(
    context: ExecutionContext,
    api_id: str,
    table_refs: list[str],
    sqlserver_target: SqlServerTarget,
    authority_root: Path | None,
) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    for table_ref in table_refs:
        target_database = sqlserver_target_database(sqlserver_target)
        schema_sql = find_schema_sql(authority_root, table_ref)
        seed_sql = find_seed_sql(authority_root, api_id, table_ref)
        if schema_sql is None or seed_sql is None:
            generated_schema_sql, generated_seed_sql = ensure_generated_authority_sql(context, api_id, table_ref, authority_root)
            schema_sql = schema_sql or generated_schema_sql
            seed_sql = seed_sql or generated_seed_sql

        actual_name = sqlserver_find_existing_table(sqlserver_target, table_ref)
        if actual_name:
            row_count = sqlserver_row_count(sqlserver_target, table_ref, actual_name)
        elif not sqlserver_database_exists(sqlserver_target, target_database):
            checks.append(
                {
                    "tableRef": table_ref,
                    "resolvedTableName": None,
                    "exists": False,
                    "rowCount": None,
                    "action": "blocked",
                    "blockReason": "database_missing",
                    "schemaSql": normalize_persisted_path(schema_sql, project_root=context.project_root),
                    "seedSql": normalize_persisted_path(seed_sql, project_root=context.project_root),
                }
            )
            continue
        else:
            row_count = None
        if actual_name and row_count and row_count > 0:
            action = "reuse"
            block_reason = None
        elif actual_name:
            action = "seed"
            block_reason = None if seed_sql else "seed_file_missing"
        else:
            action = "create"
            block_reason = None if schema_sql else MISSING_SCHEMA_AUTHORITY_REASON
        checks.append(
            {
                "tableRef": table_ref,
                "resolvedTableName": actual_name,
                "exists": bool(actual_name),
                "rowCount": row_count,
                "action": action,
                "blockReason": block_reason,
                "schemaSql": normalize_persisted_path(schema_sql, project_root=context.project_root),
                "seedSql": normalize_persisted_path(seed_sql, project_root=context.project_root),
            }
        )
    return checks


def resolve_artifact_path(project_root: Path, relative_value: str | None) -> Path | None:
    text = clean_text(relative_value)
    if not text:
        return None
    path = Path(text)
    if not path.is_absolute():
        path = (project_root / path).resolve()
    return path if path.exists() else None


def execute_sqlserver_apply(
    context: ExecutionContext,
    table_checks: list[dict[str, Any]],
    sqlserver_target: SqlServerTarget,
    *,
    allow_create_database: bool,
    allow_create_table: bool,
    allow_seed: bool,
) -> tuple[list[dict[str, Any]], list[str]]:
    executed_sql_texts: list[str] = []
    updated_checks = json.loads(json.dumps(table_checks, ensure_ascii=False))

    for check in updated_checks:
        table_ref = clean_text(check.get("tableRef"))
        target_database = sqlserver_target_database(sqlserver_target)
        schema_sql_path = resolve_artifact_path(context.project_root, check.get("schemaSql"))
        seed_sql_path = resolve_artifact_path(context.project_root, check.get("seedSql"))

        if not sqlserver_database_exists(sqlserver_target, target_database):
            if not allow_create_database:
                check["blockReason"] = "database_missing"
                continue
            sqlserver_create_database(sqlserver_target, target_database)
            executed_sql_texts.append(f"CREATE DATABASE [{target_database}];")
        actual_name = sqlserver_find_existing_table(sqlserver_target, table_ref)
        if actual_name is None:
            if not allow_create_table:
                check["blockReason"] = "table_missing_and_create_not_allowed"
                continue
            if schema_sql_path is None:
                check["blockReason"] = MISSING_SCHEMA_AUTHORITY_REASON
                continue
            sql_text = read_text(schema_sql_path)
            result = run_sqlcmd(sqlserver_target, sql_text, database_override=target_database)
            if result.returncode != 0:
                check["blockReason"] = "create_table_failed"
                continue
            executed_sql_texts.append(sql_text.strip())
            actual_name = sqlserver_find_existing_table(sqlserver_target, table_ref)
            if actual_name is None:
                check["blockReason"] = "create_table_no_effect"
                continue

        row_count = sqlserver_row_count(sqlserver_target, table_ref, actual_name)
        if row_count == 0:
            if not allow_seed:
                check["blockReason"] = "seed_required_but_not_allowed"
                check["resolvedTableName"] = actual_name
                check["exists"] = True
                check["rowCount"] = row_count
                continue
            if seed_sql_path is None:
                check["blockReason"] = "seed_file_missing"
                check["resolvedTableName"] = actual_name
                check["exists"] = True
                check["rowCount"] = row_count
                continue
            sql_text = read_text(seed_sql_path)
            result = run_sqlcmd(sqlserver_target, sql_text, database_override=target_database)
            if result.returncode != 0:
                check["blockReason"] = "seed_failed"
                check["resolvedTableName"] = actual_name
                check["exists"] = True
                check["rowCount"] = row_count
                continue
            executed_sql_texts.append(sql_text.strip())
            row_count = sqlserver_row_count(sqlserver_target, table_ref, actual_name)
            if row_count == 0:
                check["blockReason"] = "seed_no_effect"
                check["resolvedTableName"] = actual_name
                check["exists"] = True
                check["rowCount"] = row_count
                continue

        check["resolvedTableName"] = actual_name
        check["exists"] = True
        check["rowCount"] = row_count
        check["blockReason"] = None
        check["action"] = "reuse"
    return updated_checks, executed_sql_texts


def build_db_fixture_report(
    context: ExecutionContext,
    item: dict[str, Any],
    detection: dict[str, Any],
    table_checks: list[dict[str, Any]],
    *,
    status: str,
    phase: str,
    message: str,
    db_target: str | None,
    authority_root: Path | None,
    execution_mode: str,
) -> dict[str, Any]:
    blocking_tables = [check for check in table_checks if clean_text(check.get("blockReason"))]
    return {
        "schemaVersion": STATE_SCHEMA_VERSION,
        "skillName": SKILL_NAME,
        "manifestType": "db-fixture-report",
        "updatedAt": now_iso(),
        "executionId": context.execution_id,
        "apiId": item["apiId"],
        "status": status,
        "phase": phase,
        "executionMode": execution_mode,
        "dbTarget": db_target,
        "authorityRoot": normalize_persisted_path(authority_root, project_root=context.project_root),
        "sqlFixtureRequired": bool(detection.get("sqlFixtureRequired")),
        "sourcesUsed": list(detection.get("sourcesUsed") or []),
        "tableRefs": list(detection.get("tableRefs") or []),
        "summary": {
            "tableCount": len(table_checks),
            "blockingTableCount": len(blocking_tables),
            "queryTextCount": len(detection.get("queryTexts") or []),
        },
        "message": message,
        "findings": [
            {
                "tableRef": clean_text(check.get("tableRef")),
                "action": clean_text(check.get("action")),
                "blockReason": clean_text(check.get("blockReason")) or None,
            }
            for check in table_checks
        ],
    }


def build_table_checks_payload(
    context: ExecutionContext,
    item: dict[str, Any],
    table_checks: list[dict[str, Any]],
    *,
    db_target: str | None,
    authority_root: Path | None,
) -> dict[str, Any]:
    return {
        "schemaVersion": STATE_SCHEMA_VERSION,
        "skillName": SKILL_NAME,
        "manifestType": "table-checks",
        "updatedAt": now_iso(),
        "executionId": context.execution_id,
        "apiId": item["apiId"],
        "dbTarget": db_target,
        "authorityRoot": normalize_persisted_path(authority_root, project_root=context.project_root),
        "tables": table_checks,
    }


def db_target_for_report(raw_arg: str | None, db_target: SqlServerTarget | BlockedDbTarget | None) -> str | None:
    if isinstance(db_target, BlockedDbTarget):
        return db_target.raw_target
    if raw_arg:
        return redact_db_target_text(raw_arg)
    if isinstance(db_target, SqlServerTarget):
        return configured_sqlserver_target_report_label(db_target)
    return None


def build_seed_manifest(
    context: ExecutionContext,
    item: dict[str, Any],
    table_checks: list[dict[str, Any]],
    executed_sql_texts: list[str],
) -> dict[str, Any]:
    return {
        "schemaVersion": STATE_SCHEMA_VERSION,
        "skillName": SKILL_NAME,
        "manifestType": "seed-manifest",
        "updatedAt": now_iso(),
        "executionId": context.execution_id,
        "apiId": item["apiId"],
        "executedStatementCount": len([text for text in executed_sql_texts if clean_text(text)]),
        "seededTables": [
            {
                "tableRef": clean_text(check.get("tableRef")),
                "resolvedTableName": clean_text(check.get("resolvedTableName")) or None,
                "rowCount": check.get("rowCount"),
            }
            for check in table_checks
        ],
    }


def summarize_fixture_status(items: list[dict[str, Any]]) -> dict[str, int]:
    counts = {
        "total": len(items),
        "pending": 0,
        "in_progress": 0,
        "done": 0,
        "skipped": 0,
        "not_required": 0,
        "blocked": 0,
        "error": 0,
    }
    for item in items:
        status = clean_text(item.get("fixtureStatus"))
        if status in counts:
            counts[status] += 1
    return counts


def initialize_fixture_items(checklist_payload: dict[str, Any], api_map: dict[str, UpstreamApiRecord]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for raw_item in list(checklist_payload.get("items") or []):
        if not isinstance(raw_item, dict):
            continue
        api_id = clean_text(raw_item.get("apiId"))
        upstream = api_map.get(api_id)
        if upstream is None:
            continue
        source_fingerprint = clean_text(raw_item.get("specSourceFingerprint") or upstream.manifest_payload.get("specSourceFingerprint")) or None
        previous_fingerprint = clean_text(raw_item.get("fixtureSourceFingerprint"))
        same_source = bool(previous_fingerprint and previous_fingerprint == source_fingerprint)
        previous_status = clean_text(raw_item.get("fixtureStatus"))
        if upstream.status != UPSTREAM_READY_STATUS:
            fixture_status = "pending"
            fixture_phase = "waiting_spec"
            block_reason = upstream.block_reason or f"Spec API status is {upstream.status}."
        elif same_source and previous_status in FIXTURE_STATUSES:
            fixture_status = previous_status
            fixture_phase = clean_text(raw_item.get("fixturePhase")) or "pending"
            block_reason = clean_text(raw_item.get("fixtureBlockReason")) or None
        else:
            fixture_status = "pending"
            fixture_phase = "pending"
            block_reason = None
        normalized.append(
            {
                **raw_item,
                "fixtureStatus": fixture_status,
                "fixturePhase": fixture_phase,
                "fixtureBlockReason": block_reason,
                "fixtureSourceFingerprint": source_fingerprint,
            }
        )
    return normalized


def select_target_item(items: list[dict[str, Any]], api_id: str | None) -> dict[str, Any]:
    if api_id:
        for item in items:
            if clean_text(item.get("apiId")) == api_id:
                return item
        raise SkillError(f"apiId not found in checklist: {api_id}")
    for item in items:
        if clean_text(item.get("specStatus")) == UPSTREAM_READY_STATUS and clean_text(item.get("fixtureStatus")) == "pending":
            return item
    raise SkillError("未找到可执行的 fixture API。")


def update_checklist_payload(checklist_payload: dict[str, Any], items: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        **checklist_payload,
        "updatedAt": now_iso(),
        "items": items,
    }


def update_execution_state_payload(
    execution_state: dict[str, Any],
    items: list[dict[str, Any]],
    *,
    current_api_id: str | None,
    message: str,
    phase: str,
) -> dict[str, Any]:
    return {
        **execution_state,
        "updatedAt": now_iso(),
        "fixtureStatus": "blocked" if any(clean_text(item.get("fixtureStatus")) == "blocked" for item in items) else (
            "done" if all(clean_text(item.get("fixtureStatus")) in TERMINAL_FIXTURE_STATUSES for item in items if clean_text(item.get("specStatus")) == UPSTREAM_READY_STATUS)
            else "running" if any(clean_text(item.get("fixtureStatus")) == "in_progress" for item in items)
            else "waiting_fixture"
        ),
        "fixturePhase": phase,
        "fixtureUpdatedAt": now_iso(),
        "fixtureCurrentApiId": current_api_id,
        "fixtureSummary": summarize_fixture_status(items),
        "fixtureLastMessage": message,
    }


def persist_outputs(
    context: ExecutionContext,
    execution_state: dict[str, Any],
    checklist_payload: dict[str, Any],
    item: dict[str, Any],
    upstream_api: UpstreamApiRecord,
    *,
    report_payload: dict[str, Any],
    table_checks_payload: dict[str, Any],
    seed_plan_text: str | None,
    seed_executed_text: str | None,
    seed_manifest_payload: dict[str, Any] | None,
    execution_message: str,
) -> None:
    api_id = item["apiId"]
    dump_json(context.paths.db_fixture_report_path(api_id), report_payload)
    dump_json(context.paths.table_checks_path(api_id), table_checks_payload)
    if seed_plan_text is not None:
        dump_text(context.paths.seed_plan_path(api_id), seed_plan_text)
    else:
        remove_file(context.paths.seed_plan_path(api_id))
    if seed_executed_text is not None:
        dump_text(context.paths.seed_executed_path(api_id), seed_executed_text)
    else:
        remove_file(context.paths.seed_executed_path(api_id))
    if seed_manifest_payload is not None:
        dump_json(context.paths.seed_manifest_path(api_id), seed_manifest_payload)
    else:
        remove_file(context.paths.seed_manifest_path(api_id))

    manifest_payload = require_json_object("manifest.json", upstream_api.manifest_path)
    manifest_payload.update(
        {
            "fixtureStatus": item["fixtureStatus"],
            "fixturePhase": item["fixturePhase"],
            "fixtureUpdatedAt": now_iso(),
            "fixtureBlockReason": item.get("fixtureBlockReason"),
            "fixtureSourceFingerprint": item.get("fixtureSourceFingerprint"),
            "fixtureArtifacts": {
                "dbFixtureReport": normalize_persisted_path(context.paths.db_fixture_report_path(api_id), project_root=context.project_root),
                "tableChecks": normalize_persisted_path(context.paths.table_checks_path(api_id), project_root=context.project_root),
                "seedPlan": normalize_persisted_path(context.paths.seed_plan_path(api_id), project_root=context.project_root)
                if context.paths.seed_plan_path(api_id).exists()
                else None,
                "seedExecuted": normalize_persisted_path(context.paths.seed_executed_path(api_id), project_root=context.project_root)
                if context.paths.seed_executed_path(api_id).exists()
                else None,
                "seedManifest": normalize_persisted_path(context.paths.seed_manifest_path(api_id), project_root=context.project_root)
                if context.paths.seed_manifest_path(api_id).exists()
                else None,
            },
            "lastMessage": execution_message,
        }
    )
    dump_json(upstream_api.manifest_path, manifest_payload)
    dump_json(context.paths.checklist_path, checklist_payload)
    dump_json(context.paths.execution_state_path, execution_state)
    update_chain_status(
        agent_root=context.agent_dir,
        function_code=context.execution_id,
        stage="fixture",
        status=execution_state.get("fixtureStatus"),
        phase=execution_state.get("fixturePhase"),
        message=execution_state.get("fixtureLastMessage"),
        project_root=context.project_root,
        artifacts={
            "executionState": normalize_persisted_path(context.paths.execution_state_path, project_root=context.agent_dir),
            "apiChecklist": normalize_persisted_path(context.paths.checklist_path, project_root=context.agent_dir),
            "fixtureProgress": normalize_persisted_path(context.paths.progress_path, project_root=context.agent_dir),
            "currentApiId": api_id,
        },
    )
    append_progress(context.paths.progress_path, execution_message)


def main() -> int:
    configure_stdio()
    try:
        args = parse_args()
        mode = "prepare" if args.execution_mode == "auto" else args.execution_mode
        context, execution_state, checklist_payload, api_map = build_context(args)
        items = initialize_fixture_items(checklist_payload, api_map)
        selected_item = select_target_item(items, args.api_id)
        upstream_api = api_map[selected_item["apiId"]]
        if upstream_api.status != UPSTREAM_READY_STATUS or upstream_api.api_spec_payload is None:
            raise SkillError(f"{selected_item['apiId']} upstream spec is not ready.", status="blocked", diagnosis_type="upstream_not_ready")

        detection = detect_sql_fixture_need(upstream_api.api_spec_payload)
        authority_root = find_authority_root(context.project_root, args.schema_authority_root)
        db_target = parse_db_target(
            args.db_target,
            context.project_root,
            agent_dir=context.agent_dir,
            preferred_connection_names=detection.get("backendSystems") or [],
        )

        if not detection["sqlFixtureRequired"]:
            fixture_status = "skipped"
            fixture_phase = "no_sql"
            block_reason = None
            table_checks: list[dict[str, Any]] = []
            executed_sql_texts: list[str] = []
            message = f"{selected_item['apiId']} fixture skipped: no SQL dependency."
        else:
            if isinstance(db_target, BlockedDbTarget):
                table_checks = build_blocked_target_table_checks(
                    context,
                    selected_item["apiId"],
                    detection["tableRefs"],
                    db_target,
                    authority_root,
                )
            elif isinstance(db_target, SqlServerTarget):
                table_checks = inspect_sqlserver_tables(context, selected_item["apiId"], detection["tableRefs"], db_target, authority_root)
            else:
                table_checks = build_blocked_target_table_checks(
                    context,
                    selected_item["apiId"],
                    detection["tableRefs"],
                    BlockedDbTarget(
                        raw_target=".agent/config/sql-fixture-targets.local.json",
                        block_reason=MISSING_DB_TARGET_REASON,
                        message="SQL fixture target config is missing.",
                    ),
                    authority_root,
                )
            blocking_reasons = [clean_text(check.get("blockReason")) for check in table_checks if clean_text(check.get("blockReason"))]
            executed_sql_texts = []
            if (
                mode == "apply"
                and db_target is not None
                and not isinstance(db_target, BlockedDbTarget)
                and not any(reason == "missing_db_target" for reason in blocking_reasons)
            ):
                if isinstance(db_target, SqlServerTarget):
                    table_checks, executed_sql_texts = execute_sqlserver_apply(
                        context,
                        table_checks,
                        db_target,
                        allow_create_database=args.allow_create_database,
                        allow_create_table=args.allow_create_table or db_target.allow_create_table,
                        allow_seed=args.allow_seed or db_target.allow_seed,
                    )
                blocking_reasons = [clean_text(check.get("blockReason")) for check in table_checks if clean_text(check.get("blockReason"))]

            if blocking_reasons:
                fixture_status = "blocked"
                fixture_phase = "inspection_blocked" if mode == "prepare" else "apply_blocked"
                block_reason = ", ".join(dedupe_strings(blocking_reasons))
                message = f"{selected_item['apiId']} fixture blocked: {block_reason}"
            elif mode == "prepare":
                fixture_status = "pending"
                fixture_phase = "prepared"
                block_reason = None
                message = f"{selected_item['apiId']} fixture prepared; apply can continue."
            else:
                fixture_status = "done"
                fixture_phase = "applied"
                block_reason = None
                message = f"{selected_item['apiId']} fixture done."

        selected_item["fixtureStatus"] = fixture_status
        selected_item["fixturePhase"] = fixture_phase
        selected_item["fixtureBlockReason"] = block_reason
        for index, item in enumerate(items):
            if clean_text(item.get("apiId")) == selected_item["apiId"]:
                items[index] = selected_item
                break

        report_payload = build_db_fixture_report(
            context,
            selected_item,
            detection,
            table_checks,
            status=fixture_status,
            phase=fixture_phase,
            message=message,
            db_target=db_target_for_report(args.db_target, db_target),
            authority_root=authority_root,
            execution_mode=mode,
        )
        table_checks_payload = build_table_checks_payload(
            context,
            selected_item,
            table_checks,
            db_target=db_target_for_report(args.db_target, db_target),
            authority_root=authority_root,
        )
        seed_manifest_payload = build_seed_manifest(context, selected_item, table_checks, executed_sql_texts) if executed_sql_texts else None
        validate_payload_against_schema(report_payload, "db-fixture-report.schema.json", "db-fixture-report")
        validate_payload_against_schema(table_checks_payload, "table-checks.schema.json", "table-checks")
        if seed_manifest_payload is not None:
            validate_payload_against_schema(seed_manifest_payload, "seed-manifest.schema.json", "seed-manifest")

        checklist_payload = update_checklist_payload(checklist_payload, items)
        execution_state = update_execution_state_payload(
            execution_state,
            items,
            current_api_id=selected_item["apiId"],
            message=message,
            phase=fixture_phase,
        )
        seed_plan_lines: list[str] = []
        for check in table_checks:
            if clean_text(check.get("schemaSql")):
                seed_plan_lines.append(f"-- schema: {clean_text(check['schemaSql'])}")
            if clean_text(check.get("seedSql")):
                seed_plan_lines.append(f"-- seed: {clean_text(check['seedSql'])}")
        seed_plan_text = "\n".join(seed_plan_lines) + ("\n" if seed_plan_lines else "")
        seed_executed_text = "\n\n".join(text for text in executed_sql_texts if clean_text(text)) + ("\n" if executed_sql_texts else "")
        persist_outputs(
            context,
            execution_state,
            checklist_payload,
            selected_item,
            upstream_api,
            report_payload=report_payload,
            table_checks_payload=table_checks_payload,
            seed_plan_text=seed_plan_text or None,
            seed_executed_text=seed_executed_text or None,
            seed_manifest_payload=seed_manifest_payload,
            execution_message=message,
        )
        save_batch_file(context.batch_file, {**load_batch_file(context.batch_file), "activeFunctionCode": context.execution_id}, updated_by=SKILL_NAME)
        print(message)
        return 0 if fixture_status in TERMINAL_FIXTURE_STATUSES else 1
    except SkillError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
