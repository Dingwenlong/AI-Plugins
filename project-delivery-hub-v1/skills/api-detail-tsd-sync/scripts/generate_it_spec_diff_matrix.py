#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ENDPOINT_RE = re.compile(r"/ws/[A-Za-z0-9_./-]+")
BACKEND_CODE_RE = re.compile(r"\b(?:ED|ID|EC)\d{4}\b", re.IGNORECASE)
DATATEXT_RE = re.compile(r"\bDataText\d*\b", re.IGNORECASE)


def clean_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def configure_stdio() -> None:
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except OSError:
            pass


def dump_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def resolve_path(value: str | None, *, cwd: Path) -> Path | None:
    text = clean_text(value)
    if not text:
        return None
    path = Path(text).expanduser()
    if not path.is_absolute():
        path = cwd / path
    return path.resolve()


def resolve_it_spec_path(project_root: Path, agent_root: Path, function_code: str, explicit: str | None) -> Path:
    explicit_path = resolve_path(explicit, cwd=project_root)
    if explicit_path is not None:
        if not explicit_path.exists():
            raise SystemExit(f"IT Spec file not found: {explicit_path.as_posix()}")
        return explicit_path

    handoff_path = agent_root / "functions" / function_code / "handoff" / "development-handoff.json"
    if handoff_path.exists():
        payload = load_json(handoff_path)
        for item in list(payload.get("sourceFiles") or []):
            if not isinstance(item, dict):
                continue
            if clean_text(item.get("kind")).casefold() != "it-spec":
                continue
            source = resolve_path(clean_text(item.get("sourcePath")), cwd=project_root)
            if source is not None and source.exists():
                return source
            copied = resolve_path(clean_text(item.get("copiedRelativePath")), cwd=agent_root)
            if copied is not None and copied.exists():
                return copied

    compact = function_code.replace(".", "")
    candidates = []
    for path in project_root.rglob("*.xlsx"):
        name_key = path.name.casefold().replace(".", "")
        if compact.casefold() in name_key and "it spec" in path.name.casefold():
            candidates.append(path)
    if candidates:
        return sorted(candidates, key=lambda path: (path.stat().st_mtime, len(path.as_posix())), reverse=True)[0]
    raise SystemExit(f"no Customer IT SPEC found for {function_code}")


def extract_workbook_evidence(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_summaries: list[dict[str, Any]] = []
    all_lines: list[str] = []
    for sheet in workbook.worksheets:
        non_empty_lines: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [clean_text(value).replace("\n", " / ") for value in row if clean_text(value)]
            if cells:
                line = " | ".join(cells)
                non_empty_lines.append(line)
                all_lines.append(line)
        sheet_summaries.append(
            {
                "name": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "sampleLines": non_empty_lines[:12],
            }
        )

    joined = "\n".join(all_lines)
    return {
        "sheetNames": workbook.sheetnames,
        "sheets": sheet_summaries,
        "endpoints": sorted(set(ENDPOINT_RE.findall(joined))),
        "backendCodes": sorted({value.upper() for value in BACKEND_CODE_RE.findall(joined)}),
        "dataTextAliases": sorted({value for value in DATATEXT_RE.findall(joined)}, key=str.casefold)[:40],
        "hasTtdname": "Ttdname" in joined or "TTDNAME" in joined.upper(),
        "lineCount": len(all_lines),
    }


def build_d001001_d002001_items(evidence: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "id": "ITD-D001001-001",
            "area": "账号/币别取得",
            "customerSpecEvidence": "Customer IT SPEC `流程圖`/`程式邏輯`：`ws_cddebitcurrency` 取得客户臺外幣帳戶，`ws_cddebitaccount` 依币别取得账户资讯；旧筛选逻辑含 B10200。",
            "currentDesignEvidence": "现行设计以 `Exchange/GetTransDebitAccount` 取得客户可用转出/扣款账号，并以 `CommonUtil/GetCENCurr` / `CommonFunc.GetCENCurrFunc` 处理币别资料。",
            "differenceType": "旧 endpoint 拆换为新系统支撑 API",
            "impact": "medium",
            "decision": "adopt-current-design",
            "decisionReason": "旧 `ws_*` endpoint 属旧大户页面流程；新系统交付已将账号取得与币别资料拆到 Exchange/Common contract，旧筛选条件只保留为来源线索。",
            "readinessImpact": "resolved",
            "followUp": "第 02 步 API Spec 需在 legacyEvidence 记录 `ws_cddebitcurrency/ws_cddebitaccount` 来源别名。",
        },
        {
            "id": "ITD-D001001-002",
            "area": "定存查询",
            "customerSpecEvidence": "Customer IT SPEC：`ws_querytd` 取得存单资料，response 使用 `DataText0`~`DataText26` 这类旧大户展示字段。",
            "currentDesignEvidence": "现行设计使用 `Deposit/GetFixedDepositDetail`，response 以 `fixedDepositCertificates`、`fixedDepositCertificateNumber`、`linkedAccountNumber`、`fixedDepositName`、`fixedDepositAmount` 等 canonical 字段表达。",
            "differenceType": "旧展示字段转换为新系统 canonical contract",
            "impact": "high",
            "decision": "adopt-current-design",
            "decisionReason": "`DataText*` 无业务语义且不可作为新系统 contract；现行 API Detail 已定义开发可消费字段，旧字段仅作 alias/migration evidence。",
            "readinessImpact": "resolved",
            "followUp": "API Spec 写入时不得把 `DataText*` 作为对外字段；必要时记录字段来源映射。",
        },
        {
            "id": "ITD-D001001-003",
            "area": "计息明细",
            "customerSpecEvidence": "Customer IT SPEC `流程圖` 标记 `計息明細(ID0016)`；旧流程将 `ED0016` 用于提列息及已付利息总和。",
            "currentDesignEvidence": "现行 `GetFixedDepositInterestDetail` 使用 `ED0005` 查应付/累计中利息、`ED0009` 查已领利息；`ED0016` 仅保留在 `GetFixedDepositDetail` 总览/列表利息摘要来源。",
            "differenceType": "后端电文职责重新切分",
            "impact": "high",
            "decision": "adopt-current-design",
            "decisionReason": "Deposit API Detail 已明确 `GetFixedDepositInterestDetail` 不再调用 `ED0016`，以 `ED0005/ED0009` 提供明细；`ED0016` 仍服务摘要字段，避免明细 API 重复返回总览摘要。",
            "readinessImpact": "resolved",
            "followUp": "若客户坚持 IT Spec 旧口径，需回到 API Detail 修订；当前开发按已裁决现行设计推进。",
        },
        {
            "id": "ITD-D001001-004",
            "area": "定存名称修改",
            "customerSpecEvidence": "Customer IT SPEC `程式邏輯`：`定存名稱變更` 使用 `Ttdname: 定存名稱`，后端为 `ID0005`。",
            "currentDesignEvidence": "现行设计为 `Deposit/PatchFixedDepositTitle`，request 使用 `fixedDepositName`，后端来源仍为 `IRIS ID0005`。",
            "differenceType": "旧字段名改为新系统语义字段",
            "impact": "medium",
            "decision": "adopt-current-design",
            "decisionReason": "后端来源一致；`Ttdname` 是旧系统字段名，对外 contract 使用 `fixedDepositName` 更清楚，长度与空白校验由 Enterprise 二次确认。",
            "readinessImpact": "resolved",
            "followUp": "在 API Spec mappingRules 中保留 `Ttdname -> fixedDepositName` 来源映射。",
        },
        {
            "id": "ITD-D001001-005",
            "area": "历史/取消定存",
            "customerSpecEvidence": "Customer IT SPEC：已到期、已解约仍呼叫 `ED0012`；已取消资料呼叫 `ED0017` 作为预约定存取消来源。",
            "currentDesignEvidence": "现行 `GetFixedDepositDetail` 通过 `fixedDepositQueryScope` / status filter 分流，`ED0012` 负责定存归户/历史/单笔，`ED0017` 补预约/取消资料。",
            "differenceType": "旧流程合并为 scope-driven API",
            "impact": "medium",
            "decision": "adopt-current-design",
            "decisionReason": "后端电文职责与 IT Spec 一致，但对外 API 不新增历史/取消专用 endpoint；由同一 API 通过 scope 和 status filter 表达。",
            "readinessImpact": "resolved",
            "followUp": "第 02 步需在 queryContracts/logicFlow 中保留 `ED0012/ED0017` 分流说明。",
        },
    ]


def build_generic_items(function_code: str, evidence: dict[str, Any]) -> list[dict[str, Any]]:
    symbols = evidence.get("endpoints") or evidence.get("backendCodes") or []
    symbol_text = ", ".join(symbols[:12]) if symbols else "未抽出明确旧 endpoint/backend code"
    return [
        {
            "id": f"ITD-{function_code.replace('.', '')}-001",
            "area": "Customer IT SPEC 总体差异",
            "customerSpecEvidence": f"Customer IT SPEC 已读取；抽出线索：{symbol_text}。",
            "currentDesignEvidence": "尚未自动比对到现行 TSD/API Detail 的逐项设计口径。",
            "differenceType": "待人工比对",
            "impact": "high",
            "decision": "needs-customer-confirmation",
            "decisionReason": "自动脚本只能抽取旧系统证据，尚未完成逐项裁决。",
            "readinessImpact": "blocking",
            "followUp": "补齐逐项差异矩阵后重新物化 handoff。",
        }
    ]


def build_matrix(function_code: str, it_spec_path: Path) -> dict[str, Any]:
    evidence = extract_workbook_evidence(it_spec_path)
    key = function_code.upper()
    if key == "D.001.001_D.002.001":
        items = build_d001001_d002001_items(evidence)
    else:
        items = build_generic_items(function_code, evidence)
    return {
        "schemaVersion": "1.0.0",
        "artifactType": "it-spec-diff-matrix",
        "functionCode": function_code,
        "generatedAt": dt.datetime.now(dt.timezone.utc).astimezone().isoformat(timespec="seconds"),
        "sourceWorkbook": it_spec_path.as_posix(),
        "workbookEvidence": evidence,
        "items": items,
    }


def render_markdown(matrix: dict[str, Any]) -> str:
    lines = [
        f"# {matrix['functionCode']} Customer IT SPEC 差异矩阵",
        "",
        f"- Source: `{matrix['sourceWorkbook']}`",
        f"- Generated: `{matrix['generatedAt']}`",
        "",
        "| ID | 范围 | Customer IT SPEC 口径 | 现行 TSD/API Detail 口径 | 差异类型 | 影响 | 裁决 | Ready 影响 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for item in matrix["items"]:
        lines.append(
            "| {id} | {area} | {customer} | {current} | {difference} | {impact} | {decision}：{reason} | {ready} |".format(
                id=item["id"],
                area=item["area"],
                customer=item["customerSpecEvidence"].replace("|", "\\|"),
                current=item["currentDesignEvidence"].replace("|", "\\|"),
                difference=item["differenceType"],
                impact=item["impact"],
                decision=item["decision"],
                reason=item["decisionReason"].replace("|", "\\|"),
                ready=item["readinessImpact"],
            )
        )
    lines.extend(
        [
            "",
            "## Follow-up",
            "",
        ]
    )
    for item in matrix["items"]:
        lines.append(f"- `{item['id']}` {item['followUp']}")
    lines.append("")
    return "\n".join(lines)


def main() -> int:
    configure_stdio()
    parser = argparse.ArgumentParser(description="Generate Customer IT SPEC difference matrix artifacts.")
    parser.add_argument("--project-root", default=str(Path.cwd()))
    parser.add_argument("--agent-root", default=".agent")
    parser.add_argument("--function-code", required=True)
    parser.add_argument("--it-spec")
    parser.add_argument("--output-dir")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    project_root = resolve_path(args.project_root, cwd=Path.cwd()) or Path.cwd()
    agent_root = resolve_path(args.agent_root, cwd=project_root) or (project_root / ".agent")
    function_code = clean_text(args.function_code)
    it_spec_path = resolve_it_spec_path(project_root, agent_root, function_code, args.it_spec)
    output_dir = resolve_path(args.output_dir, cwd=project_root) or (agent_root / "functions" / function_code / "analysis")

    matrix = build_matrix(function_code, it_spec_path)
    markdown = render_markdown(matrix)
    json_path = output_dir / "it-spec-diff-matrix.json"
    md_path = output_dir / "it-spec-diff-matrix.md"

    if not args.dry_run:
        output_dir.mkdir(parents=True, exist_ok=True)
        dump_json(json_path, matrix)
        md_path.write_text(markdown, encoding="utf-8")

    print(
        json.dumps(
            {
                "status": "ok",
                "functionCode": function_code,
                "itSpec": it_spec_path.as_posix(),
                "markdown": md_path.as_posix(),
                "json": json_path.as_posix(),
                "itemCount": len(matrix["items"]),
                "blockingCount": sum(1 for item in matrix["items"] if clean_text(item.get("readinessImpact")).casefold() == "blocking"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
