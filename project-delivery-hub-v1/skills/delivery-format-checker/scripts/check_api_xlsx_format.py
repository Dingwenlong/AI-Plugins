#!/usr/bin/env python3
"""Check project API Detail workbook formatting without writing the workbook."""

from __future__ import annotations

import argparse
import base64
import json
import re
import subprocess
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from project_rules import resolve_asset_path


CATEGORIES = ("Must fix", "Should fix", "Naming", "Visual risk", "Covered")
API_REQUIRED_SECTIONS = (
    "API  Name",
    "Request",
    "Response",
    "範例",
    "For中台開發人員",
    "API 內部業務邏輯",
)
TITLE_MERGES = {
    "Request": "A:{row}:G:{row}",
    "Response": "A:{row}:G:{row}",
    "範例": "A:{row}:F:{row}",
    "For中台開發人員": "A:{row}:F:{row}",
    "API 內部業務邏輯": "A:{row}:F:{row}",
}
INTERNAL_LOGIC_MERGE = "B{row}:F{row}"
EXAMPLE_REQUEST_MERGE = "B{row}:C{row}"
EXAMPLE_RESPONSE_MERGE = "D{row}:F{row}"
INTERNAL_LINK_RE = re.compile(r"^#?'?(?P<sheet>[^']+)'?!A1$", re.IGNORECASE)
PREFERRED_TERMS = {
    "校驗": "驗證 / 檢核",
    "校验": "驗證 / 檢核",
}


def configure_utf8_stdio() -> None:
    """Keep Windows console output stable for mixed Simplified/Traditional Chinese."""
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")


@dataclass
class Finding:
    category: str
    code: str
    sheet: str
    location: str
    message: str
    detail: str = ""


def norm(value: Any) -> str:
    return re.sub(r"\s+", "", "" if value is None else str(value).strip())


def display(value: Any) -> str:
    return "" if value is None else str(value).strip()


def add(
    findings: list[Finding],
    category: str,
    code: str,
    sheet: str,
    location: str,
    message: str,
    detail: str = "",
) -> None:
    findings.append(Finding(category, code, sheet, location, message, detail))


def check_preferred_terminology(findings: list[Finding], worksheets: Iterable[Any]) -> None:
    matches: list[str] = []
    for ws in worksheets:
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                text = display(cell.value)
                if not text:
                    continue
                for term, replacement in PREFERRED_TERMS.items():
                    start_at = 0
                    while True:
                        index = text.find(term, start_at)
                        if index < 0:
                            break
                        start = max(0, index - 12)
                        end = min(len(text), index + len(term) + 12)
                        excerpt = text[start:end]
                        matches.append(f"{ws.title}!{cell.coordinate}: `{term}` -> `{replacement}` in `{excerpt}`")
                        start_at = index + len(term)

    if matches:
        add(
            findings,
            "Should fix",
            "language.preferred_terminology",
            "Workbook",
            "visible text",
            "Use project preferred wording: replace `校驗` with `驗證` for validation meaning or `檢核` for checklist/checking meaning.",
            "; ".join(matches[:20]) + ("; ..." if len(matches) > 20 else ""),
        )
    else:
        add(findings, "Covered", "language.preferred_terminology", "Workbook", "visible text", "No disallowed `校驗` wording was found.")


def is_api_list_sheet(ws: Any) -> bool:
    if ws.title in {"Api_List", "API_List"}:
        return True
    return "PRD" in norm(ws["A1"].value) and "API" in norm(ws["E1"].value)


def row_values(ws: Any, row: int, max_col: int = 7) -> Iterable[str]:
    for col in range(1, max_col + 1):
        yield display(ws.cell(row=row, column=col).value)


def is_api_detail_sheet(ws: Any) -> bool:
    if is_api_list_sheet(ws):
        return False
    if norm(ws["A1"].value) == "APIName":
        return True

    labels = {"Request", "Response", "範例", "API內部業務邏輯"}
    found = set()
    for row in range(1, min(ws.max_row, 120) + 1):
        for text in row_values(ws, row):
            key = norm(text)
            if key in labels:
                found.add(key)
    return len(found) >= 2


def find_section_rows(ws: Any) -> dict[str, int]:
    rows: dict[str, int] = {}
    if norm(ws["A1"].value) == "APIName":
        rows["API  Name"] = 1
    for row in range(1, ws.max_row + 1):
        # Section title labels are left-anchored in column A. Do not treat field
        # table headers or scenario labels in B:G as section boundaries.
        text = norm(ws.cell(row=row, column=1).value)
        for label in API_REQUIRED_SECTIONS[1:]:
            if text == norm(label) and label not in rows:
                rows[label] = row
    return rows


def semantic_last_row(ws: Any, max_col: int = 7) -> int:
    for row in range(ws.max_row, 0, -1):
        for col in range(1, max_col + 1):
            if display(ws.cell(row=row, column=col).value):
                return row
    return 0


def has_border(cell: Any) -> bool:
    border = cell.border
    return any(side is not None and side.style for side in (border.left, border.right, border.top, border.bottom))


def has_right_border(cell: Any) -> bool:
    side = cell.border.right
    return side is not None and bool(side.style)


def has_bottom_border(cell: Any) -> bool:
    side = cell.border.bottom
    return side is not None and bool(side.style)


def has_top_border(cell: Any) -> bool:
    side = cell.border.top
    return side is not None and bool(side.style)


def has_fill(cell: Any) -> bool:
    fill = cell.fill
    return bool(fill and fill.fill_type)


def fill_rgb(cell: Any) -> str:
    color = cell.fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    return ""


def font_rgb(cell: Any) -> str:
    color = cell.font.color
    if color is not None and color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    return ""


def fill_matches_config(cell: Any, fill_config: dict[str, Any] | None) -> bool:
    if fill_config is None:
        return not has_fill(cell)
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False
    expected = fill_config.get("fgColor", {})
    actual = fill.fgColor
    if expected.get("type") == "rgb":
        return actual.type == "rgb" and (actual.rgb or "").upper() == expected.get("rgb", "").upper()
    if expected.get("type") == "theme":
        try:
            return (
                actual.type == "theme"
                and int(actual.theme) == int(expected["theme"])
                and abs(float(actual.tint or 0) - float(expected.get("tint", 0))) < 0.01
            )
        except (TypeError, ValueError):
            return False
    return False


def fill_signature(cell: Any) -> tuple[Any, ...]:
    fill = cell.fill
    if not fill or not fill.fill_type:
        return ("none",)

    color = fill.fgColor
    if color.type == "rgb":
        color_value: Any = (color.rgb or "").upper()
    elif color.type == "theme":
        try:
            color_value = (int(color.theme), round(float(color.tint or 0), 4))
        except (TypeError, ValueError):
            color_value = (str(color.theme), str(color.tint))
    elif color.type == "indexed":
        color_value = color.indexed
    else:
        color_value = str(color.type)

    return (fill.fill_type, color.type, color_value)


def alignment_matches_config(cell: Any, alignment_config: dict[str, Any]) -> bool:
    alignment = cell.alignment
    expected_horizontal = alignment_config.get("horizontal")
    expected_vertical = alignment_config.get("vertical")
    expected_wrap = alignment_config.get("wrapText")

    if expected_horizontal and alignment.horizontal != expected_horizontal:
        return False
    if expected_vertical and alignment.vertical != expected_vertical:
        return False
    if expected_wrap is not None and bool(alignment.wrap_text) != bool(expected_wrap):
        return False
    return True


def is_blank(cell: Any) -> bool:
    return not display(cell.value)


def merged_ranges(ws: Any) -> set[str]:
    return {str(rng) for rng in ws.merged_cells.ranges}


def required_merge(label: str, row: int) -> str:
    template = TITLE_MERGES[label]
    start_col, start_row_marker, end_col, end_row_marker = template.split(":")
    return f"{start_col}{row}:{end_col}{row}"


def internal_logic_merge(row: int) -> str:
    return INTERNAL_LOGIC_MERGE.format(row=row)


def example_request_merge(row: int) -> str:
    return EXAMPLE_REQUEST_MERGE.format(row=row)


def example_response_merge(row: int) -> str:
    return EXAMPLE_RESPONSE_MERGE.format(row=row)


def row_has_visible_content(ws: Any, row: int, min_col: int = 1, max_col: int = 6) -> bool:
    return any(display(ws.cell(row=row, column=col).value) for col in range(min_col, max_col + 1))


def effective_row_height(ws: Any, row: int) -> float:
    height = ws.row_dimensions[row].height
    if height is None:
        return 15.0
    return float(height)


def check_fixed_row_height(
    findings: list[Finding],
    ws: Any,
    row: int,
    expected: float,
    location: str,
    label: str,
) -> None:
    actual = effective_row_height(ws, row)
    if abs(actual - expected) > 0.6:
        add(
            findings,
            "Should fix",
            "title_height.template_fixed",
            ws.title,
            location,
            f"{label} must keep the configured template row height.",
            f"actual={actual}; expected={expected}",
        )


def check_api_name_description(findings: list[Finding], ws: Any, config: dict[str, Any]) -> None:
    region = config["regions"]["apiNameDescription"]
    header_height = float(region["headerRow"]["rowHeight"])
    content_height = float(region["contentRow"]["rowHeight"])

    check_fixed_row_height(findings, ws, 1, header_height, "A1:B1", "API Name / API Description header row")
    check_fixed_row_height(findings, ws, 2, content_height, "A2:B2", "API method / description content row")

    if region["headerRow"].get("apiDescriptionRightBorderRequired", False) and not has_right_border(ws["B1"]):
        add(
            findings,
            "Should fix",
            "api_name_description.right_border",
            ws.title,
            "B1",
            "API Description header cell must keep a right border before the C1 return link.",
        )

    if region["contentRow"].get("apiDescriptionRightBorderRequired", False) and not has_right_border(ws["B2"]):
        add(
            findings,
            "Should fix",
            "api_name_description.content_right_border",
            ws.title,
            "B2",
            "API Description content cell must keep a right border before the return-link area.",
        )

    if has_border(ws["C1"]) or has_fill(ws["C1"]):
        add(
            findings,
            "Should fix",
            "api_name_description.return_link_style",
            ws.title,
            "C1",
            "Return link cell must stay hyperlink-only: no table border and no fill.",
        )


def check_template_row_heights(findings: list[Finding], ws: Any, sections: dict[str, int], config: dict[str, Any]) -> None:
    row_heights = config["global"]["rowHeights"]
    section_height = float(row_heights["sectionTitle"])
    table_header_height = float(row_heights["tableHeader"])
    middle_office_height = float(row_heights["middleOfficeTitle"])

    for label, row in sections.items():
        if label == "API  Name":
            continue
        expected = middle_office_height if label == "For中台開發人員" else section_height
        check_fixed_row_height(findings, ws, row, expected, f"A{row}:G{row}", f"{label} section title row")

        header_row = row + 1
        if label in {"Request", "Response", "範例", "API 內部業務邏輯"} and row_has_visible_content(ws, header_row, 1, 7):
            check_fixed_row_height(findings, ws, header_row, table_header_height, f"A{header_row}:G{header_row}", f"{label} table/header row")


def check_internal_logic_merges(findings: list[Finding], ws: Any, sections: dict[str, int], merges: set[str]) -> None:
    logic_title_row = sections.get("API 內部業務邏輯")
    if not logic_title_row:
        return

    last_semantic = semantic_last_row(ws)
    if last_semantic <= logic_title_row:
        return

    checked = 0
    block_end = logic_title_row
    for row in range(logic_title_row + 1, last_semantic + 1):
        if not row_has_visible_content(ws, row, 1, 6):
            if checked:
                break
            continue
        checked += 1
        block_end = row
        expected = internal_logic_merge(row)
        if expected in merges:
            continue

        extra_values = [
            f"{get_column_letter(col)}{row}={display(ws.cell(row=row, column=col).value)}"
            for col in range(3, 7)
            if display(ws.cell(row=row, column=col).value)
        ]
        if extra_values:
            add(
                findings,
                "Should fix",
                "merge.internal_logic_row",
                ws.title,
                expected,
                "Internal logic row must merge B:F; C:F contains content that must be consolidated before merging.",
                "; ".join(extra_values),
            )
        else:
            add(
                findings,
                "Should fix",
                "merge.internal_logic_row",
                ws.title,
                expected,
                f"Internal logic row should be merged as {expected}.",
            )

    if checked:
        add(
            findings,
            "Covered",
            "merge.internal_logic_rows_checked",
            ws.title,
            f"B{logic_title_row + 1}:F{block_end}",
            f"Checked {checked} contiguous internal logic header/content rows for B:F merges.",
        )


def check_internal_logic_label_styles(findings: list[Finding], ws: Any, sections: dict[str, int], config: dict[str, Any]) -> None:
    logic_title_row = sections.get("API 內部業務邏輯")
    if not logic_title_row:
        return

    logic_config = config["regions"]["internalLogic"]["logicRows"]["labelColumn"]
    alignment_config = config["alignments"][logic_config["alignment"]]
    fill_config = config["fills"][logic_config["fill"]]
    fill_must_match = logic_config.get("fillMustMatch")
    header_fill_signature = None
    if fill_must_match == "internalLogic.tableHeaderRow.labelColumn":
        header_fill_signature = fill_signature(ws.cell(row=logic_title_row + 1, column=1))
    last_semantic = semantic_last_row(ws)
    checked = 0
    for row in range(logic_title_row + 2, last_semantic + 1):
        if not row_has_visible_content(ws, row, 1, 6):
            if checked:
                break
            continue
        checked += 1
        cell = ws.cell(row=row, column=1)
        alignment_ok = alignment_matches_config(cell, alignment_config)
        fill_ok = (
            fill_signature(cell) == header_fill_signature
            if header_fill_signature is not None
            else fill_matches_config(cell, fill_config)
        )
        if not (alignment_ok and fill_ok):
            add(
                findings,
                "Should fix",
                "style.internal_logic_label_column",
                ws.title,
                f"A{row}",
                "Internal logic label column must be left-aligned, vertically centered, wrapped, and use the same fill as the internal-logic table header row.",
                f"alignment_ok={alignment_ok}; fill_ok={fill_ok}",
            )


def merge_extra_values(ws: Any, row: int, cols: Iterable[int]) -> list[str]:
    return [
        f"{get_column_letter(col)}{row}={display(ws.cell(row=row, column=col).value)}"
        for col in cols
        if display(ws.cell(row=row, column=col).value)
    ]


def check_example_merges(findings: list[Finding], ws: Any, sections: dict[str, int], merges: set[str]) -> None:
    example_title_row = sections.get("範例")
    if not example_title_row:
        return

    next_section_rows = [
        row
        for label, row in sections.items()
        if row > example_title_row and label in {"For中台開發人員", "API 內部業務邏輯"}
    ]
    last_semantic = semantic_last_row(ws)
    section_end = min(next_section_rows) - 1 if next_section_rows else last_semantic
    if section_end <= example_title_row:
        return

    checked = 0
    for row in range(example_title_row + 1, section_end + 1):
        if not row_has_visible_content(ws, row, 1, 6):
            if checked:
                break
            continue

        checked += 1
        expected_request = example_request_merge(row)
        expected_response = example_response_merge(row)
        is_header = row == example_title_row + 1
        code = "merge.example_header_row" if is_header else "merge.example_scenario_row"
        label = "Example header row" if is_header else "Example scenario row"

        if expected_request not in merges:
            extra_values = merge_extra_values(ws, row, (3,))
            add(
                findings,
                "Should fix",
                code,
                ws.title,
                expected_request,
                f"{label} Request area should be merged as {expected_request}.",
                "; ".join(extra_values),
            )

        if expected_response not in merges:
            extra_values = merge_extra_values(ws, row, (5, 6))
            message = f"{label} Response area should be merged as {expected_response}."
            if extra_values:
                message = f"{message} E:F contains content that must be consolidated before merging."
            add(
                findings,
                "Should fix",
                code,
                ws.title,
                expected_response,
                message,
                "; ".join(extra_values),
            )

    if checked:
        add(
            findings,
            "Covered",
            "merge.example_rows_checked",
            ws.title,
            f"B{example_title_row + 1}:F{example_title_row + checked}",
            f"Checked {checked} example header/scenario rows for B:C and D:F merges.",
        )


def check_example_label_styles(findings: list[Finding], ws: Any, sections: dict[str, int], config: dict[str, Any]) -> None:
    example_title_row = sections.get("範例")
    if not example_title_row:
        return

    next_section_rows = [
        row
        for label, row in sections.items()
        if row > example_title_row and label in {"For中台開發人員", "API 內部業務邏輯"}
    ]
    last_semantic = semantic_last_row(ws)
    section_end = min(next_section_rows) - 1 if next_section_rows else last_semantic
    if section_end <= example_title_row + 1:
        return

    label_config = config["regions"]["example"]["scenarioRows"]["labelColumn"]
    alignment_config = config["alignments"][label_config["alignment"]]
    checked = 0
    for row in range(example_title_row + 2, section_end + 1):
        if not row_has_visible_content(ws, row, 1, 6):
            if checked:
                break
            continue
        checked += 1
        cell = ws.cell(row=row, column=1)
        if not alignment_matches_config(cell, alignment_config):
            add(
                findings,
                "Should fix",
                "style.example_scenario_label_alignment",
                ws.title,
                f"A{row}",
                "Example scenario label cells must be left-aligned, vertically centered, and wrapped.",
            )


def next_section_row(sections: dict[str, int], current_row: int) -> int:
    later_rows = sorted(row for row in sections.values() if row > current_row)
    return later_rows[0] if later_rows else 0


def last_populated_row_in_range(ws: Any, start_row: int, end_row: int, max_col: int) -> int:
    if start_row <= 0 or end_row < start_row:
        return 0
    for row in range(end_row, start_row - 1, -1):
        if row_has_visible_content(ws, row, 1, max_col):
            return row
    return 0


def visible_bottom_border_at(ws: Any, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if has_bottom_border(cell):
        return True
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            if row == merged_range.max_row and has_bottom_border(ws.cell(row=merged_range.min_row, column=merged_range.min_col)):
                return True
            for merged_col in range(merged_range.min_col, merged_range.max_col + 1):
                if has_bottom_border(ws.cell(row=merged_range.max_row, column=merged_col)):
                    return True
    if row < ws.max_row:
        return has_top_border(ws.cell(row=row + 1, column=col))
    return False


def check_row_bottom_closure(
    findings: list[Finding],
    ws: Any,
    row: int,
    min_col: int,
    max_col: int,
    code: str,
    label: str,
) -> bool:
    missing = [get_column_letter(col) for col in range(min_col, max_col + 1) if not visible_bottom_border_at(ws, row, col)]
    if missing:
        add(
            findings,
            "Should fix",
            code,
            ws.title,
            f"{get_column_letter(min_col)}{row}:{get_column_letter(max_col)}{row}",
            f"{label} last visible row must have a closed bottom border.",
            f"missing_bottom_columns={','.join(missing)}",
        )
        return False
    return True


def check_section_bottom_borders(findings: list[Finding], ws: Any, sections: dict[str, int]) -> None:
    checked = 0
    for label, max_col in (("Request", 7), ("Response", 7), ("範例", 6), ("API 內部業務邏輯", 6)):
        title_row = sections.get(label, 0)
        if not title_row:
            continue
        next_row = next_section_row(sections, title_row)
        end_row = (next_row - 1) if next_row else semantic_last_row(ws, max_col)
        last_row = last_populated_row_in_range(ws, title_row + 2, end_row, max_col)
        if not last_row:
            continue
        checked += 1
        check_row_bottom_closure(
            findings,
            ws,
            last_row,
            1,
            max_col,
            "api_detail.section_bottom_border",
            label,
        )
    if checked:
        add(
            findings,
            "Covered",
            "api_detail.section_bottom_border_checked",
            ws.title,
            "A:G",
            f"Checked bottom-border closure for {checked} API Detail content sections.",
        )


def internal_link_target(cell: Any) -> str:
    if not cell.hyperlink:
        return ""
    return cell.hyperlink.location or cell.hyperlink.target or ""


def link_points_to_sheet(target: str, sheet_name: str) -> bool:
    compact = target.strip()
    if not compact:
        return False
    compact = compact.lstrip("#")
    expected = f"'{sheet_name}'!A1"
    if compact.lower() == expected.lower():
        return True
    match = INTERNAL_LINK_RE.match(target.strip())
    return bool(match and match.group("sheet").lower() == sheet_name.lower())


def find_api_name_column(api_list: Any) -> int:
    for row in range(1, min(api_list.max_row, 5) + 1):
        for col in range(1, api_list.max_column + 1):
            text = norm(api_list.cell(row=row, column=col).value)
            if text in {"API名稱", "APIName"}:
                return col
    return 5


def column_width(ws: Any, col: int) -> float:
    width = ws.column_dimensions[get_column_letter(col)].width
    return float(width or 8.43)


def estimate_wrapped_row_height(ws: Any, row: int, min_col: int = 1, max_col: int = 10) -> float:
    max_lines = 1
    for col in range(min_col, max_col + 1):
        text = display(ws.cell(row=row, column=col).value)
        if not text:
            continue
        width = max(column_width(ws, col), 1.0)
        cell_lines = 0
        for line in text.splitlines() or [text]:
            # Excel character metrics vary by font; use a conservative visual
            # heuristic only to catch obvious clipped rows.
            cell_lines += max(1, int((len(line) + max(width * 1.15, 1) - 1) // max(width * 1.15, 1)))
        max_lines = max(max_lines, cell_lines)
    return min(max(20.1, max_lines * 15.0), 409.5)


def check_api_list_row_visuals(
    findings: list[Finding],
    api_list: Any,
    cell: Any,
    method: str,
    config: dict[str, Any],
) -> None:
    expected_font = config.get("fonts", {}).get("hyperlink", {})
    expected_color = expected_font.get("color", {}).get("rgb", "FF0563C1").upper()
    actual_color = font_rgb(cell)
    underline = cell.font.underline
    if not actual_color or actual_color != expected_color or str(underline).lower() not in {"single", "true"} or (cell.font.name or "") != expected_font.get("name", "Times New Roman"):
        add(
            findings,
            "Should fix",
            "api_list.hyperlink_style",
            api_list.title,
            cell.coordinate,
            "Api_List API name hyperlink must use the configured shallow-blue underlined style.",
            f"method={method}; font={cell.font.name}; color={actual_color}; underline={underline}",
        )

    horizontal = cell.alignment.horizontal or ""
    vertical = cell.alignment.vertical or ""
    if horizontal.lower() != "left" or vertical.lower() != "center":
        add(
            findings,
            "Should fix",
            "api_list.api_name_alignment",
            api_list.title,
            cell.coordinate,
            "Api_List API name cells must be left-aligned and vertically centered.",
            f"method={method}; horizontal={horizontal or '<default>'}; vertical={vertical or '<default>'}",
        )

    if not visible_bottom_border_at(api_list, cell.row, cell.column):
        add(
            findings,
            "Should fix",
            "api_list.bottom_border",
            api_list.title,
            cell.coordinate,
            "Api_List API name cell must have a visible thin black bottom border.",
            f"method={method}",
        )

    row = cell.row
    wrapped_missing = []
    for col in range(1, min(api_list.max_column, 10) + 1):
        row_cell = api_list.cell(row=row, column=col)
        if display(row_cell.value) and row_cell.alignment.wrap_text is not True:
            wrapped_missing.append(row_cell.coordinate)
    if wrapped_missing:
        add(
            findings,
            "Should fix",
            "api_list.row_autofit",
            api_list.title,
            f"A{row}:J{row}",
            "Api_List target rows must enable wrap text before Excel COM AutoFit.",
            f"method={method}; missing_wrap={','.join(wrapped_missing[:8])}",
        )

    actual_height = api_list.row_dimensions[row].height or 15.0
    estimated_height = estimate_wrapped_row_height(api_list, row, 1, min(api_list.max_column, 10))
    if actual_height + 0.25 < estimated_height:
        add(
            findings,
            "Visual risk",
            "visual.row_height_clipping",
            api_list.title,
            f"A{row}:J{row}",
            "Api_List row height may clip wrapped text; run Excel COM AutoFit and visually recheck.",
            f"method={method}; actual={actual_height}; estimated_min={estimated_height}",
        )


def check_api_list_links(findings: list[Finding], wb: Any, api_sheets: list[Any], config: dict[str, Any]) -> None:
    api_list = next((ws for ws in wb.worksheets if is_api_list_sheet(ws)), None)
    if api_list is None:
        add(findings, "Should fix", "api_list.missing", "Workbook", "", "Missing Api_List worksheet.")
        return

    api_name_col = find_api_name_column(api_list)
    by_method: dict[str, list[Any]] = {}
    for row in range(2, api_list.max_row + 1):
        cell = api_list.cell(row=row, column=api_name_col)
        method = display(cell.value)
        if method:
            by_method.setdefault(method, []).append(cell)

    checked = 0
    for ws in api_sheets:
        method = display(ws["A2"].value)
        if not method:
            continue
        checked += 1
        cells = by_method.get(method, [])
        if not cells:
            add(
                findings,
                "Should fix",
                "api_list.link_row_missing",
                "Api_List",
                f"column {get_column_letter(api_name_col)}",
                f"Api_List has no row for API method {method}.",
                ws.title,
            )
            continue
        good = any(link_points_to_sheet(internal_link_target(cell), ws.title) for cell in cells)
        if not good:
            add(
                findings,
                "Should fix",
                "api_list.link_target",
                "Api_List",
                ",".join(cell.coordinate for cell in cells),
                f"Api_List link for {method} does not point to the API worksheet A1.",
                ws.title,
            )
        for cell in cells:
            if link_points_to_sheet(internal_link_target(cell), ws.title):
                check_api_list_row_visuals(findings, api_list, cell, method, config)

    add(
        findings,
        "Covered",
        "api_list.links_checked",
        "Api_List",
        f"column {get_column_letter(api_name_col)}",
        f"Checked Api_List link coverage for {checked} API worksheets.",
    )


def check_sheet(findings: list[Finding], ws: Any, config: dict[str, Any]) -> None:
    sections = find_section_rows(ws)
    for label in API_REQUIRED_SECTIONS:
        if label not in sections:
            add(findings, "Must fix", "section.missing", ws.title, label, f"Missing required section {label}.")

    ordered = [(label, sections[label]) for label in API_REQUIRED_SECTIONS if label in sections]
    if ordered != sorted(ordered, key=lambda item: item[1]):
        add(
            findings,
            "Must fix",
            "section.order",
            ws.title,
            "A:G",
            "Required API Detail sections are not in the expected order.",
            " > ".join(f"{label}@{row}" for label, row in ordered),
        )
    elif len(ordered) == len(API_REQUIRED_SECTIONS):
        add(findings, "Covered", "section.order", ws.title, "A:G", "Required API Detail sections are present and ordered.")

    merges = merged_ranges(ws)
    for label, row in sections.items():
        if label in TITLE_MERGES:
            expected = required_merge(label, row)
            if expected not in merges:
                add(
                    findings,
                    "Should fix",
                    "merge.title",
                    ws.title,
                    f"A{row}",
                    f"Section title {label} should be merged as {expected}.",
                )

    check_api_name_description(findings, ws, config)
    check_template_row_heights(findings, ws, sections, config)
    check_example_merges(findings, ws, sections, merges)
    check_internal_logic_merges(findings, ws, sections, merges)
    check_example_label_styles(findings, ws, sections, config)
    check_internal_logic_label_styles(findings, ws, sections, config)
    check_section_bottom_borders(findings, ws, sections)

    widths = config["global"]["columnWidths"]
    for col, expected in widths.items():
        actual = ws.column_dimensions[col].width
        if actual is None:
            continue
        if abs(float(actual) - float(expected)) > 1.5:
            add(
                findings,
                "Should fix",
                "style.column_width",
                ws.title,
                col,
                f"Column {col} width differs from configured baseline.",
                f"actual={actual}; expected={expected}",
            )

    if ws.sheet_view.showGridLines is not False:
        add(findings, "Should fix", "style.gridlines", ws.title, "", "API Detail worksheet gridlines should be hidden.")

    rows = sections
    for label, expected_fill in (("Request", "FFBDD7EE"), ("Response", "FFBDD7EE"), ("範例", "FFBDD7EE"), ("API 內部業務邏輯", "FFBDD7EE"), ("For中台開發人員", "FFFFFF00")):
        row = rows.get(label)
        if not row:
            continue
        actual = fill_rgb(ws.cell(row=row, column=1))
        if actual and actual != expected_fill:
            add(
                findings,
                "Should fix",
                "style.title_fill",
                ws.title,
                f"A{row}",
                f"Section title {label} fill differs from configured color.",
                f"actual={actual}; expected={expected_fill}",
            )

    last_semantic = semantic_last_row(ws)
    used_range = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    semantic_range = f"A1:G{last_semantic}" if last_semantic else "A1:G0"
    outside_nonempty = 0
    outside_styled = 0
    outside_bordered = 0
    outside_filled = 0
    for row in range(1, ws.max_row + 1):
        for col in range(8, min(ws.max_column, 52) + 1):
            cell = ws.cell(row=row, column=col)
            if not is_blank(cell):
                outside_nonempty += 1
            elif has_border(cell) or has_fill(cell):
                outside_styled += 1
                if has_border(cell):
                    outside_bordered += 1
                if has_fill(cell):
                    outside_filled += 1

    if (ws.max_column > 7 or ws.max_row > last_semantic + 5) and (outside_nonempty or outside_filled):
        add(
            findings,
            "Visual risk",
            "scope.used_range_expanded",
            ws.title,
            used_range,
            "Excel used range extends beyond the semantic API Detail range and still has outside-range values or fills.",
            f"semantic_range={semantic_range}",
        )

    if outside_nonempty:
        add(
            findings,
            "Should fix",
            "scope.outside_visible_values",
            ws.title,
            "H:AZ",
            "Cells outside the configured visible API Detail range contain values.",
            f"count={outside_nonempty}",
        )
    if outside_filled:
        add(
            findings,
            "Visual risk",
            "scope.outside_visible_styles",
            ws.title,
            "H:AZ",
            "Blank cells outside the configured visible range have visible fills; this can create apparent table spillover.",
            f"styled={outside_styled}; bordered={outside_bordered}; filled={outside_filled}",
        )

    bottom_styled = 0
    if last_semantic:
        for row in range(last_semantic + 1, ws.max_row + 1):
            for col in range(1, min(ws.max_column, 52) + 1):
                cell = ws.cell(row=row, column=col)
                if is_blank(cell) and has_fill(cell):
                    bottom_styled += 1
    if bottom_styled:
        add(
            findings,
            "Visual risk",
            "scope.bottom_styles",
            ws.title,
            f"A{last_semantic + 1}:AZ{ws.max_row}",
            "Blank rows below the last API Detail content row still contain visible styles.",
            f"styled={bottom_styled}",
        )

    add(
        findings,
        "Covered",
        "scope.semantic_range",
        ws.title,
        semantic_range,
        "Computed semantic API Detail range for safe repair.",
        f"excel_used_range={used_range}",
    )

    if config.get("global", {}).get("rowHeights", {}).get("autoFitVisibleRows"):
        add(
            findings,
            "Covered",
            "style.row_height_autofit_policy",
            ws.title,
            f"A1:G{last_semantic}",
            "Visible API Detail rows must be auto-fit after merge repair; merged rows require Excel COM measurement.",
        )


def load_config(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def check_com_outside_visible_borders(findings: list[Finding], workbook_path: Path, api_sheets: list[Any]) -> None:
    if not api_sheets:
        return

    powershell_exe = Path(r"C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe")
    powershell_cmd = str(powershell_exe) if powershell_exe.exists() else "powershell.exe"
    payload = {
        "path": str(workbook_path.resolve()),
        "sheets": [{"name": ws.title, "lastRow": semantic_last_row(ws)} for ws in api_sheets],
    }
    encoded = base64.b64encode(json.dumps(payload, ensure_ascii=False).encode("utf-8")).decode("ascii")
    script = rf"""
$ErrorActionPreference = 'Stop'
$payloadJson = [Text.Encoding]::UTF8.GetString([Convert]::FromBase64String('{encoded}'))
$payload = $payloadJson | ConvertFrom-Json
$xlLineStyleNone = -4142
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$workbook = $null
$results = @()
try {{
  $workbook = $excel.Workbooks.Open($payload.path, $null, $true)
  foreach ($sheetInfo in $payload.sheets) {{
    $worksheet = $workbook.Worksheets.Item([string]$sheetInfo.name)
    $lastRow = [int]$sheetInfo.lastRow
    $bordered = 0
    $sample = @()
    if ($lastRow -gt 0) {{
      $outsideRange = $worksheet.Range($worksheet.Cells.Item(1, 8), $worksheet.Cells.Item($lastRow, 52))
      $rangeHasBorder = $false
      foreach ($borderIndex in @(8, 9, 11, 12)) {{
        $lineStyle = $outsideRange.Borders.Item($borderIndex).LineStyle
        if ($lineStyle -ne $xlLineStyleNone) {{
          $rangeHasBorder = $true
          break
        }}
      }}
      if ($rangeHasBorder) {{
        $bordered = 1
        for ($row = 1; $row -le [Math]::Min($lastRow, 80); $row++) {{
          for ($col = 8; $col -le 52; $col++) {{
            $cell = $worksheet.Cells.Item($row, $col)
            $hasCellBorder = $false
            foreach ($borderIndex in @(8, 9, 11, 12)) {{
              if ($cell.Borders.Item($borderIndex).LineStyle -ne $xlLineStyleNone) {{
                $hasCellBorder = $true
                break
              }}
            }}
            if (-not $hasCellBorder -and $col -gt 8 -and $cell.Borders.Item(7).LineStyle -ne $xlLineStyleNone) {{
              $hasCellBorder = $true
            }}
            if (-not $hasCellBorder -and $col -lt 52 -and $cell.Borders.Item(10).LineStyle -ne $xlLineStyleNone) {{
              $hasCellBorder = $true
            }}
            if ($hasCellBorder) {{
              $sample += $cell.Address($false, $false)
              if ($sample.Count -ge 5) {{
                break
              }}
            }}
          }}
          if ($sample.Count -ge 5) {{
            break
          }}
        }}
      }}
    }}
    $results += [PSCustomObject]@{{
      Sheet = [string]$sheetInfo.name
      Bordered = $bordered
      Sample = ($sample -join ',')
    }}
  }}
  $results | ConvertTo-Json -Compress
}} finally {{
  if ($null -ne $workbook) {{
    $workbook.Close($false) | Out-Null
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
  }}
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
  [System.GC]::Collect()
  [System.GC]::WaitForPendingFinalizers()
}}
"""
    try:
        encoded_script = base64.b64encode(script.encode("utf-16le")).decode("ascii")
        completed = subprocess.run(
            [powershell_cmd, "-NoProfile", "-ExecutionPolicy", "Bypass", "-EncodedCommand", encoded_script],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=120,
            check=False,
        )
    except Exception as exc:
        add(
            findings,
            "Visual risk",
            "scope.outside_visible_borders_com_unavailable",
            "Workbook",
            "H:AZ",
            "Excel COM visible border check could not run; render or COM inspection is still required for outside-range borders.",
            str(exc),
        )
        return

    if completed.returncode != 0:
        add(
            findings,
            "Visual risk",
            "scope.outside_visible_borders_com_unavailable",
            "Workbook",
            "H:AZ",
            "Excel COM visible border check failed; render or COM inspection is still required for outside-range borders.",
            completed.stderr.strip(),
        )
        return

    text = completed.stdout.strip()
    if not text:
        return

    try:
        records = json.loads(text)
    except json.JSONDecodeError as exc:
        add(
            findings,
            "Visual risk",
            "scope.outside_visible_borders_com_unavailable",
            "Workbook",
            "H:AZ",
            "Excel COM visible border check returned non-JSON output.",
            str(exc),
        )
        return

    if isinstance(records, dict):
        records = [records]

    checked = 0
    for record in records:
        checked += 1
        bordered = int(record.get("Bordered", 0))
        if bordered:
            add(
                findings,
                "Visual risk",
                "scope.outside_visible_borders_com",
                record.get("Sheet", ""),
                "H:AZ",
                "Blank cells outside the visible API Detail range have visible Excel COM borders.",
                f"bordered={bordered}; sample={record.get('Sample', '')}",
            )

    add(
        findings,
        "Covered",
        "scope.outside_visible_borders_com_checked",
        "Workbook",
        "H:AZ",
        f"Checked visible Excel COM borders outside A:G for {checked} API worksheets.",
    )


def text_report(findings: list[Finding]) -> str:
    lines: list[str] = []
    for category in CATEGORIES:
        items = [finding for finding in findings if finding.category == category]
        lines.append(f"{category}: {len(items)}")
        for item in items:
            location = f" [{item.sheet}{'!' + item.location if item.location else ''}]"
            detail = f" ({item.detail})" if item.detail else ""
            lines.append(f"- {item.code}{location}: {item.message}{detail}")
    return "\n".join(lines)


def main() -> int:
    configure_utf8_stdio()
    parser = argparse.ArgumentParser(description="Check project API Detail workbook format without saving.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--config", type=Path, default=None)
    parser.add_argument("--rules-root", help="专案规则库根目录；未传入时读取 workspace 配置。")
    parser.add_argument("--sheets", nargs="*", help="Optional explicit API worksheet names to check.")
    parser.add_argument("--format", choices=("text", "json"), default="text")
    args = parser.parse_args()

    config_path = args.config or resolve_asset_path(
        "apiDetailExcelStyle",
        rules_root_arg=args.rules_root,
    )
    if config_path is None:
        raise SystemExit("project API Detail Excel style config not found; pass --rules-root or --config.")
    config = load_config(config_path)
    wb = load_workbook(args.workbook, read_only=False, data_only=False)
    findings: list[Finding] = []

    if args.sheets:
        api_sheets = [wb[name] for name in args.sheets]
    else:
        api_sheets = [ws for ws in wb.worksheets if is_api_detail_sheet(ws)]

    if not api_sheets:
        add(findings, "Must fix", "workbook.no_api_sheets", "Workbook", "", "No API Detail worksheets were detected.")
    else:
        add(findings, "Covered", "workbook.api_sheet_detection", "Workbook", "", f"Detected {len(api_sheets)} API Detail worksheets.")

    for ws in api_sheets:
        check_sheet(findings, ws, config)

    check_preferred_terminology(findings, wb.worksheets)
    check_com_outside_visible_borders(findings, args.workbook, api_sheets)
    check_api_list_links(findings, wb, api_sheets, config)

    if args.format == "json":
        print(json.dumps([asdict(finding) for finding in findings], ensure_ascii=False, indent=2))
    else:
        print(text_report(findings))

    return 1 if any(f.category == "Must fix" for f in findings) else 0


if __name__ == "__main__":
    raise SystemExit(main())
